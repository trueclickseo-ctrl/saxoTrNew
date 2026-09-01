"""
Phase 2 of reports/live_pair_commission_analysis.py -- turns
data/live_pair_commission_analysis.json into a formatted .xlsx with two
sheets:
  1. "LIVE pair economics" -- per-pair cost + net-after-cost at EUR45 risk
  2. "Saxo pricing reference" -- the facts read off the two Saxo
     rates-and-conditions pages on 2026-09-01

    py -3.12 reports/live_pair_commission_analysis_xlsx.py
"""
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(BASE, "data", "live_pair_commission_analysis.json")
OUT = os.path.join(BASE, "data", "live_pair_commission_analysis.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
TITLE_FONT = Font(bold=True, size=13)
NOTE_FONT = Font(italic=True, size=9, color="555555")

COLS = [
    ("pair", "Pair", 10),
    ("tier", "Tier", 16),
    ("price", "Price", 10),
    ("spread_pips", "Live spread\n(pips)", 10),
    ("spread_pct_of_price", "Spread\n% of price", 10),
    ("spread_cost_eur_roundtrip", "Spread cost €\n(round trip)", 12),
    ("commission_eur_roundtrip", "Saxo commission €\n(round trip)", 14),
    ("commission_pct_of_notional", "Commission\n% of notional", 12),
    ("min_ticket_binds", "1-USD min\nticket binds?", 11),
    ("tomnext_markup_pct", "Tom/Next\nmarkup %/yr", 11),
    ("tomnext_eur_per_day", "Financing €\n/ day held", 11),
    ("units_at_E45", "Units\n@ €45 risk", 11),
    ("notional_eur_at_E45", "Notional €\n@ €45 risk", 12),
    ("realised_risk_eur", "Realised\nrisk € (R)", 10),
    ("all_in_cost_eur", "All-in cost €\n(comm+sprd+slip)", 13),
    ("recovery_0p5R_eur", "0.5R recovery\n€", 11),
    ("recovery_to_cost_ratio", "0.5R ÷ all-in\n(gate ≥ 3.0)", 12),
    ("rejected_at_e45_risk", "Rejected\n@ €45 risk?", 11),
    ("tp_2R_net_after_comm_eur", "TP (2R) net\nafter cost €", 12),
    ("bounce_0p5R_net_after_comm_eur", "0.5R bounce net\nafter comm €", 13),
    ("verdict", "Verdict", 46),
]

REFERENCE = [
    ("SOURCE PAGES", ""),
    ("Spreads & commissions", "https://www.home.saxo/rates-and-conditions/forex/spreads-and-commissions"),
    ("Trading conditions", "https://www.home.saxo/rates-and-conditions/forex/trading-conditions"),
    ("Read on", "2026-09-01"),
    ("", ""),
    ("FX SPOT / FORWARD PRICING (published retail plan)", ""),
    ("Commission on FX spot/forward", "NONE on the published plan - priced entirely through the bid/ask spread."),
    ("Pricing tiers", "Classic / Platinum / VIP  ('pay less as you trade more')"),
    ("EURUSD indicative spread", "Classic 1.0 pip  |  Platinum 0.9 pip  |  VIP 0.7 pip"),
    ("Minimum commission (transaction fee)", "1 USD per side, charged on small-notional trades. Round trip => 2 USD floor."),
    ("Minimum trade size", "per-pair, shown under the info button on the forex pricing page"),
    ("", ""),
    ("THIS ACCOUNT (what Saxo actually charges us - live API + real fills)", ""),
    ("FX spot commission on OUR account", "FLAT ~EUR2.59 per side  (~EUR5.18 round trip), pair- and size-independent at our sizes."),
    ("Confirmed by", "live /trade/v1/infoprices Commissions.CostBuy for every pair  AND  the real closed MXNUSD position (Cost -5.19)."),
    ("Implication", "Our account is on a commission schedule, not the spread-only retail plan. The flat fee - not the spread - is the cost that matters."),
    ("Spread cost at EUR45 sizing", "only ~EUR0.4-1.3 round trip (see per-pair sheet) - an order of magnitude smaller than the flat commission."),
    ("", ""),
    ("OVERNIGHT FINANCING (Tom/Next markup) - a HOLDING cost, not an entry cost", ""),
    ("Classic tier markup", "+/- 0.75% per year on position value"),
    ("Platinum / VIP", "+/- 0.60%  /  +/- 0.50%"),
    ("Extra markup on MXN, RUB, TRY, ZAR crosses", "+/- 0.30%  (=> 1.05% Classic on those pairs)"),
    ("Financing interest", "daily market overnight rate +/- 2.00% markup"),
    ("", ""),
    ("TRADING HOURS", ""),
    ("FX", "24 h/day, 5.5 days/week. Mon 05:04 Sydney -> Fri 16:59 EST."),
    ("Non-tradeable window", "prices stop streaming 1 min before 17:00 EST and stay grey for 5 min - do not trade then."),
    ("Special-hours currencies", "AED, ILS, RON, THB, TRY have restricted daily windows."),
    ("Precious metals (XAU/XAG/XPT)", "18:01 to 16:59 EST"),
    ("", ""),
    ("HOW ATOS USES THIS", ""),
    ("LIVE cost gate (2026-09-01, user)", "blocks a trade when (a) round-trip cost / FX rate lookup fails, (b) expected 2R target profit < 5x round-trip commission, or (c) RSI: a 0.5R recovery < 3.0x the all-in transaction cost."),
    ("Recovery-vs-cost gate", "ONE pair-independent rule. all_in_cost = flat commission + one spread crossing + 0.5-pip slippage. REPLACED both MIN_LIVE_NOTIONAL_EUR and the pair-specific LIVE_RSI_MIN_UNITS table."),
    ("Reject, never resize up", "if realised R is too low for a 0.5R recovery to clear 3x cost, the SIGNAL IS REJECTED. EUR45 is a risk CEILING; bumping size up to fix the ratio would breach it."),
    ("ASSUMED_EXIT_R = 0.5", "provisional stand-in for the real median RSI(2) exit. The AI trade journal will measure it over ~1 week of clean LIVE data; then this one constant is updated (mechanism unchanged)."),
    ("Why pair-independent", "at fixed EUR45 risk, realised R is EUR37-45 and commission is a flat EUR5.18 for EVERY pair => 0.5R clears all-in cost by 3.1-3.9x on all 17. There is no per-pair number to encode."),
    ("MXNUSD lesson", "that loss was a LEGACY 1,000-unit trade, notional ~EUR1,090, R collapsed to ~EUR5 => 0.5R (EUR2.5) < 3x cost. Rejected by this gate. The EUR45 fixed-risk sizing prevents R collapse in the first place."),
]


def _econ_sheet(wb, data):
    rows = data["rows"]
    ws = wb.active
    ws.title = "LIVE pair economics"

    ws["A1"] = "ATOS LIVE - per-pair commission economics at EUR45 fixed RSI risk"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"Generated {data['generated'][:19]}  ·  RISK_EUR={data['risk_eur']}  ·  "
                f"TP={data['tp_rr']}R  ·  gate: {data['assumed_exit_r']}R recovery ≥ "
                f"{data['min_recovery_mult']}× all-in cost  (pair-independent, reject-not-resize)")
    ws["A2"].font = NOTE_FONT
    ws["A3"] = ("Saxo charges THIS account a FLAT ~EUR5.18 round-trip commission on FX spot (not a spread - "
                "confirmed live + by the real MXNUSD fill). Spread + slippage add only ~EUR0.6-1.6. At the "
                "fixed EUR45 risk every pair's economics are the same, so the LIVE gate is ONE rule: a "
                f"{data['assumed_exit_r']}R recovery must clear {data['min_recovery_mult']}x the all-in cost. "
                "If realised R collapses (tight stop + lot rounding) the signal is REJECTED, never resized up. "
                f"ASSUMED_EXIT_R={data['assumed_exit_r']} is provisional - the AI journal will measure the real "
                "median RSI exit and this one constant gets updated.")
    ws["A3"].font = NOTE_FONT

    hr = 5
    for c, (_, label, width) in enumerate(COLS, 1):
        cell = ws.cell(row=hr, column=c, value=label)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[hr].height = 30
    ws.freeze_panes = f"A{hr+1}"

    for i, r in enumerate(rows):
        rr = hr + 1 + i
        for c, (key, _, _) in enumerate(COLS, 1):
            v = r.get(key)
            if isinstance(v, bool):
                v = "YES" if v else "no"
            ws.cell(row=rr, column=c, value=v)
        v = (r.get("verdict") or "")
        fill = (OK_FILL if v.startswith("OK") else
                WARN_FILL if v.startswith("legacy") else BAD_FILL)
        ws.cell(row=rr, column=len(COLS)).fill = fill

    sr = hr + 2 + len(rows)
    hv = [r for r in rows if not str(r.get("verdict", "")).startswith("legacy")]
    rejected = [r for r in hv if r.get("rejected_at_e45_risk")]
    ratios = [r["recovery_to_cost_ratio"] for r in hv if r.get("recovery_to_cost_ratio")]
    ws.cell(row=sr, column=1, value="LIVE RSI universe:").font = Font(bold=True)
    ws.cell(row=sr, column=2, value=len(hv))
    ws.cell(row=sr, column=3, value=(f"{len(hv) - len(rejected)} clear the gate at €45 risk, {len(rejected)} rejected  ·  "
                                     f"0.5R÷all-in ratio range {min(ratios):.2f}–{max(ratios):.2f} (threshold {data['min_recovery_mult']})"))
    ws.cell(row=sr+1, column=1, value="Rejected at €45 risk:").font = Font(bold=True)
    ws.cell(row=sr+1, column=3, value=", ".join(r["pair"] for r in rejected) or "none — all 17 clear it")
    ws.cell(row=sr+2, column=1, value="Note:").font = Font(bold=True)
    ws.cell(row=sr+2, column=3, value=("The gate only bites if realised R collapses. GBP crosses sit closest to the "
                                       "threshold (~3.1×) — if their ATR widens they could flip. The MXNUSD-class "
                                       "failure (R ≈ €5) is fully blocked."))


def _ref_sheet(wb):
    ws = wb.create_sheet("Saxo pricing reference")
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 110
    ws["A1"] = "Saxo FX pricing - as published on the two rates-and-conditions pages (2026-09-01)"
    ws["A1"].font = TITLE_FONT
    r = 3
    for k, v in REFERENCE:
        a = ws.cell(row=r, column=1, value=k)
        b = ws.cell(row=r, column=2, value=v)
        if k and not v:                       # section header
            a.font = Font(bold=True, size=11, color="1F4E78")
        elif k:
            a.font = Font(bold=True, size=10)
            b.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1


def main():
    data = json.load(open(SRC, encoding="utf-8"))
    wb = openpyxl.Workbook()
    _econ_sheet(wb, data)
    _ref_sheet(wb)
    wb.save(OUT)
    rej = [r for r in data["rows"] if r.get("rejected_at_e45_risk")]
    print(f"wrote {OUT}  ({len(data['rows'])} pairs, {len(rej)} rejected at €45 risk: "
          f"{', '.join(r['pair'] for r in rej) or '-'})")


if __name__ == "__main__":
    main()
