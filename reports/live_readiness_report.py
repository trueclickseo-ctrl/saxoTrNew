"""
reports/live_readiness_report.py
----------------------------------
Strategy x Pair LIVE readiness matrix for the 17-pair HIGH_VOLUME_SYMBOLS
universe and the 3 approved LIVE strategies (bb, rsi, pullback).

Per-cell verdict logic (a starting framework, not a magic threshold --
revisit the exact cutoffs once a real forward sample exists):
  NO DATA               -- zero trades
  INSUFFICIENT SAMPLE   -- fewer than 5 trades (not enough to judge)
  LIVE candidate        -- net-positive AND commission < 30% of |gross|
  SIM only (cost heavy) -- net-positive but cost eats too much of the edge
  DO NOT TRADE LIVE     -- net-negative

Two-phase design, same reason as daily_sim_report.py (forex.runner
transitively imports torch, which isn't on whichever Python has
openpyxl in this environment):

    python reports/_gather_daily_sim_data.py     # phase 1: real Saxo data
    py -3.12 reports/live_readiness_report.py     # phase 2: build the .xlsx

Read-only analysis only -- never touches a live signal, gate, stop, or
order. Safe to run any time, including during the forward-SIM
"no-touch" observation period.

IMPORTANT CAVEAT, as of the first run (2026-08-27/28): only 7 of the 51
strategy x pair cells have ANY historical data, and none have the 5+
trades needed for the stats to mean anything -- because this data
predates the 17-pair narrowing (SIM was trading a 149-pair, 10-strategy
universe when these trades happened). This is NOT yet the fresh
forward-SIM sample the LIVE decision needs (50-100 qualifying trades
across bb/rsi/pullback on these 17 pairs specifically) -- re-run this
once that sample has accumulated, don't treat an early run's near-empty
matrix as a real answer.
"""
import sys, os, json
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
from forex.universe import HIGH_VOLUME_SYMBOLS

_trades_cache_path = os.path.join(BASE_DIR, ".devtools", "_daily_sim_trades.json")
if not os.path.exists(_trades_cache_path):
    print(f"ERROR: {_trades_cache_path} not found.")
    print("Run phase 1 first:  python reports/_gather_daily_sim_data.py")
    sys.exit(1)
with open(_trades_cache_path) as f:
    trades = json.load(f)

LIVE_STRATS = ["bb", "rsi", "pullback"]
usable = [t for t in trades if t["commission_eur"] is not None and t["gross_pnl_eur"] is not None]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill("solid", fgColor="FCE4E4")
GREEN_FILL = PatternFill("solid", fgColor="E4F7E4")
YELLOW_FILL = PatternFill("solid", fgColor="FFF6D9")
GREY_FILL = PatternFill("solid", fgColor="EEEEEE")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, headers, row=1):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDER


def compute_stats(ts):
    n = len(ts)
    gross = sum(t["gross_pnl_eur"] for t in ts)
    comm = sum(t["commission_eur"] for t in ts)
    net = sum(t["net_pnl_eur"] for t in ts)
    wins = [t for t in ts if t["net_pnl_eur"] > 0]
    losses = [t for t in ts if t["net_pnl_eur"] <= 0]
    win_rate = round(len(wins) / n * 100, 1) if n else None
    gross_win_sum = sum(t["net_pnl_eur"] for t in wins)
    gross_loss_sum = abs(sum(t["net_pnl_eur"] for t in losses))
    pf = round(gross_win_sum / gross_loss_sum, 2) if gross_loss_sum > 0 else (float("inf") if gross_win_sum > 0 else None)
    expectancy = round(net / n, 2) if n else None
    cost_pct_gross = round(comm / abs(gross) * 100, 1) if gross != 0 else None
    avg_cost = round(comm / n, 2) if n else None
    return {"n": n, "gross": round(gross, 2), "comm": round(comm, 2), "net": round(net, 2),
            "win_rate": win_rate, "pf": pf, "expectancy": expectancy,
            "cost_pct_gross": cost_pct_gross, "avg_cost": avg_cost}


def verdict(stats):
    if stats["n"] == 0:
        return "NO DATA"
    if stats["n"] < 5:
        return "INSUFFICIENT SAMPLE"
    if stats["net"] > 0 and stats["cost_pct_gross"] is not None and stats["cost_pct_gross"] < 30:
        return "LIVE candidate"
    if stats["net"] > 0:
        return "SIM only (cost heavy)"
    return "DO NOT TRADE LIVE"


VERDICT_FILL = {
    "NO DATA": GREY_FILL, "INSUFFICIENT SAMPLE": YELLOW_FILL,
    "LIVE candidate": GREEN_FILL, "SIM only (cost heavy)": YELLOW_FILL,
    "DO NOT TRADE LIVE": RED_FILL,
}

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: 17-Pair x 3-Strategy Readiness Matrix
# ============================================================
ws = wb.active
ws.title = "17-Pair Readiness Matrix"
ws["A1"] = "LIVE Readiness -- 17 High-Volume Pairs x bb/rsi/pullback"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = (f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} PKT. Uses SIM's existing historical trades "
            f"filtered to strategy-in-{{bb,rsi,pullback}} AND pair-in-17-high-volume. If most cells show 0-2 "
            f"trades, that's expected until enough real forward-SIM data accumulates under the narrowed "
            f"configuration -- don't treat a near-empty matrix as a real go/no-go answer.")
ws["A2"].font = Font(italic=True, color="B91C1C", bold=True)
ws.merge_cells("A2:L2")
ws["A2"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[2].height = 60

style_header(ws, ["Strategy", "Pair", "Trades", "Gross P&L (EUR)", "Commission (EUR)",
                   "Net P&L (EUR)", "Win Rate %", "Profit Factor", "Expectancy/Trade (EUR)",
                   "Cost % of |Gross|", "Avg Cost/Trade (EUR)", "Verdict"], row=4)
row = 5
by_cell = defaultdict(list)
for t in usable:
    if t["strategy"] in LIVE_STRATS and t["symbol"] in HIGH_VOLUME_SYMBOLS:
        by_cell[(t["strategy"], t["symbol"])].append(t)

for strat in LIVE_STRATS:
    for sym in sorted(HIGH_VOLUME_SYMBOLS):
        ts = by_cell.get((strat, sym), [])
        s = compute_stats(ts)
        v = verdict(s)
        vals = [strat, sym, s["n"], s["gross"], s["comm"], s["net"], s["win_rate"],
                s["pf"], s["expectancy"], s["cost_pct_gross"], s["avg_cost"], v]
        for i, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=i, value=val)
            c.border = BORDER
            if i == 12:
                c.fill = VERDICT_FILL.get(v, GREY_FILL)
                c.font = BOLD
        row += 1
LAST = row - 1
for i, w in enumerate([10, 9, 8, 14, 14, 14, 10, 12, 16, 14, 14, 20], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

n_no_data = sum(1 for r in range(5, LAST + 1) if ws.cell(row=r, column=3).value == 0)
n_total = LAST - 4

# ============================================================
# Sheet 2: Strategy Overview, ALL pairs (broader context, more data)
# ============================================================
ws2 = wb.create_sheet("Strategy Overview (All Pairs)")
ws2["A1"] = "Broader context: each approved strategy across ALL pairs it has ever traded in SIM"
ws2["A1"].font = Font(bold=True, size=13)
ws2["A2"] = ("Not the 17-pair-specific answer -- more data points (helps judge the strategy's own character) "
             "but includes pairs that will never be part of the LIVE 17-pair universe.")
ws2["A2"].font = Font(italic=True, color="666666")
ws2.merge_cells("A2:J2")
style_header(ws2, ["Strategy", "Trades", "Gross P&L (EUR)", "Commission (EUR)", "Net P&L (EUR)",
                    "Win Rate %", "Profit Factor", "Expectancy/Trade (EUR)", "Cost % of |Gross|",
                    "Verdict"], row=4)
row = 5
for strat in LIVE_STRATS:
    ts = [t for t in usable if t["strategy"] == strat]
    s = compute_stats(ts)
    v = verdict(s)
    vals = [strat, s["n"], s["gross"], s["comm"], s["net"], s["win_rate"], s["pf"],
            s["expectancy"], s["cost_pct_gross"], v]
    for i, val in enumerate(vals, 1):
        c = ws2.cell(row=row, column=i, value=val)
        c.border = BORDER
        if i == 10:
            c.fill = VERDICT_FILL.get(v, GREY_FILL)
            c.font = BOLD
    row += 1
for i, w in enumerate([12, 8, 15, 15, 15, 10, 12, 17, 14, 20], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 3: Cost definition notes
# ============================================================
ws3 = wb.create_sheet("Cost Definition Notes")
notes = [
    ("What Saxo's 'Cost' actually is",
     "Confirmed live via Saxo's own API (both /trade/v1/infoprices Commissions and "
     "/port/v1/closedpositions/me CostOpening/CostClosing): this is COMMISSION only -- "
     "a distinct line item, separate from price-based P&L. It is NOT spread, and for a "
     "same-day round trip it is NOT financing either."),
    ("Where spread actually is",
     "Saxo does not expose spread as a separate EUR figure at all. It's embedded in the real "
     "executed fill price itself (you buy at the ask, sell at the bid) -- which IS what our "
     "entry_price/exit_price already are (fetched fresh at execution time, not a stale signal "
     "price). Verified directly: comparing stored entry/exit against Saxo's own real "
     "OpenPrice/ClosingPrice for actual closed trades showed differences of well under 1 pip "
     "for real-time-logged closes -- consistent with normal quote timing, not a systematically "
     "missing cost. Spread is therefore already reflected in 'Gross P&L' above, not a hidden "
     "extra column to add."),
    ("Where financing/swap shows up",
     "Bundled into the same Cost figure once a position is held overnight -- confirmed earlier "
     "by comparing a same-day close's Cost against a multi-day-held position's Cost for the "
     "same pair/size: the multi-day one was measurably larger. It is NOT separable from "
     "commission in Saxo's own API response -- 'Total Cost' in these tables already includes "
     "whatever accrued financing applied, it just can't be split into a distinct Financing "
     "column without a dedicated multi-day-holding-period study."),
    ("Practical conclusion for the tables in this workbook",
     "'Commission (EUR)' = real, live-verified, separate from Gross. 'Spread' is not a "
     "meaningful separate column to report -- it's already inside Gross P&L. 'Financing' is "
     "inside 'Commission (EUR)' for any multi-day-held trade, not broken out. Net P&L = Gross - "
     "Commission is therefore already Net P&L = Gross - ALL actual costs, to the precision "
     "Saxo's own API exposes."),
]
style_header(ws3, ["Topic", "Finding"])
for i, (topic, finding) in enumerate(notes, 2):
    ws3.cell(row=i, column=1, value=topic).font = BOLD
    c = ws3.cell(row=i, column=2, value=finding)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws3.row_dimensions[i].height = 90
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 110

# ============================================================
# Sheet 0: Summary (first sheet)
# ============================================================
ws0 = wb.create_sheet("Summary", 0)
ws0["A1"] = "LIVE Readiness Report -- Summary"
ws0["A1"].font = Font(bold=True, size=14)
ws0["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} PKT"
ws0["A2"].font = Font(italic=True, color="666666")

r0 = 4
ws0.cell(row=r0, column=1, value="17-pair x 3-strategy cells with ANY historical data").font = BOLD
ws0.cell(row=r0, column=2, value=f"{n_total - n_no_data} of {n_total}")
r0 += 1
ws0.cell(row=r0, column=1, value="Cells with >=5 trades (minimum for the stats to mean anything)").font = BOLD
cells_5plus = sum(1 for (s, sym), ts in by_cell.items() if len(ts) >= 5)
ws0.cell(row=r0, column=2, value=cells_5plus)
r0 += 2
ws0.cell(row=r0, column=1, value="VERDICT").font = Font(bold=True, size=13, color="B91C1C")
r0 += 1
ws0.cell(row=r0, column=1, value=(
    "Not enough forward-SIM data yet to fill the 17-pair readiness matrix meaningfully -- "
    f"only {sum(len(v) for v in by_cell.values())} historical trades exist across the entire "
    "17x3 grid, because SIM was trading a much broader universe when they happened. The report "
    "structure and verdict logic are built and ready (see sheet 1) -- what's missing is the "
    "fresh qualifying sample (50-100 trades across bb/rsi/pullback on these 17 pairs "
    "specifically) this report needs to actually answer the question. See 'Strategy Overview' "
    "for the broader (still thin) context in the meantime."
)).alignment = Alignment(wrap_text=True)
ws0.merge_cells(f"A{r0}:F{r0+3}")
ws0.column_dimensions["A"].width = 55
ws0.column_dimensions["B"].width = 20

out_path = os.path.join(os.path.expanduser("~"), "Downloads", "live_readiness_report_17pairs.xlsx")
wb.save(out_path)
print("Saved:", out_path)
print(f"Cells with data: {n_total - n_no_data}/{n_total}, cells with >=5 trades: {cells_5plus}")
