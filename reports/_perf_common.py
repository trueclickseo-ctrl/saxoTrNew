"""
reports/_perf_common.py
-------------------------
Shared helpers for the per-module performance trackers (forex's
reports/pair_group_performance_tracker.py and the generalized
reports/module_performance_tracker.py for futures/etf/stock) --
2026-08-28. Both build a persistent workbook with the same underlying
shape: a hidden "Trade Detail" sheet, plus summary sheets that group its
rows via live COUNTIF/SUMIF/SUMIFS formulas rather than pre-computed
static values, so re-running the build always reflects the current
Trade Detail contents even if a sheet is manually copied/inspected.

Common Trade Detail column layout every summary-sheet formula assumes:
  A Strategy | B Symbol | C Group | D Direction | E Units | F Status
  G Gross P&L | H Commission | I Net P&L | J Net Result (WIN/LOSS)
  K Close Date (YYYY-MM-DD, blank if still open)
  L Week (ISO, "YYYY-Www", blank if still open)
  M Month ("YYYY-MM", blank if still open)
Column C (Group) is forex-only (its 8 Forex Grouping tiers) -- the
generalized module tracker leaves it blank; every formula here only
ever references A/B/G/H/I/J/K/L/M, never C, so this is safe either way.
"""
from datetime import date


def parse_close_date(ts):
    """Best-effort parse of a timestamp_close string into a date().
    Returns None for missing/unparseable/still-open trades."""
    if not ts:
        return None
    s = str(ts)[:10]
    try:
        return date.fromisoformat(s)
    except Exception:
        return None


def day_key(d: date) -> str:
    return d.isoformat()


def week_key(d: date) -> str:
    iso = d.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def month_key(d: date) -> str:
    return f"{d.year:04d}-{d.month:02d}"


def sorted_distinct_period_keys(trades: list, key_fn) -> list:
    """Distinct period keys (day/week/month) present among CLOSED trades
    with a parseable close date, sorted ascending (oldest first, matching
    a normal top-to-bottom "how did we do over time" reading order).
    Open positions have no close date and are excluded from every
    time-bucketed sheet -- they don't yet have a "when this happened" to
    bucket by; they're already visible in the Per-Strategy/Per-Symbol
    sheets' "open" tracking where applicable."""
    keys = set()
    for t in trades:
        if t.get("status") != "closed":
            continue
        d = parse_close_date(t.get("timestamp_close"))
        if d is None:
            continue
        keys.add(key_fn(d))
    return sorted(keys)


def write_metric_formulas(ws, row_i: int, crit: str, td_name: str, last_row: int, start_col: int):
    """Writes Trades/Wins/WR%/Gross/Commission/Net/Profit Factor as 7 live
    formulas into columns [start_col .. start_col+6] of row `row_i`,
    filtered by `crit` (an already-built COUNTIF/SUMIF-style criteria
    fragment, e.g. "'Trade Detail'!$C$2:$C$99,\"Metals\"" or
    "'Trade Detail'!$K$2:$K$99,\"2026-08-27\"").

    `td_name` is the (already-quoted) Trade Detail sheet reference, e.g.
    "'Trade Detail'". `last_row` is Trade Detail's last data row.
    """
    from openpyxl.utils import get_column_letter as _L
    c = start_col
    trades_col, wins_col, wr_col, gross_col, comm_col, net_col, pf_col = (
        _L(c), _L(c + 1), _L(c + 2), _L(c + 3), _L(c + 4), _L(c + 5), _L(c + 6)
    )
    ws.cell(row=row_i, column=c,     value=f'=COUNTIF({crit})')
    ws.cell(row=row_i, column=c + 1, value=f'=COUNTIFS({crit},{td_name}!$J$2:$J${last_row},"WIN")')
    ws.cell(row=row_i, column=c + 2, value=f'=IFERROR(ROUND({wins_col}{row_i}/{trades_col}{row_i}*100,1),"")')
    ws.cell(row=row_i, column=c + 3, value=f'=ROUND(SUMIF({crit},{td_name}!$G$2:$G${last_row}),2)')
    ws.cell(row=row_i, column=c + 4, value=f'=ROUND(SUMIF({crit},{td_name}!$H$2:$H${last_row}),2)')
    ws.cell(row=row_i, column=c + 5, value=f'=ROUND(SUMIF({crit},{td_name}!$I$2:$I${last_row}),2)')
    win_sum = f'SUMIFS({td_name}!$I$2:$I${last_row},{crit},{td_name}!$I$2:$I${last_row},">0")'
    loss_sum = f'SUMIFS({td_name}!$I$2:$I${last_row},{crit},{td_name}!$I$2:$I${last_row},"<0")'
    ws.cell(row=row_i, column=c + 6,
            value=f'=IFERROR(ROUND({win_sum}/ABS({loss_sum}),2),IF({trades_col}{row_i}=0,"",">0 (no losers)"))')


METRIC_HEADERS = ["Trades", "Wins (net)", "WR %", "Gross P&L", "Commission", "Net P&L", "Profit Factor"]
