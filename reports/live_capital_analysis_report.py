"""
reports/live_capital_analysis_report.py
------------------------------------------
Phase 2 (build the .xlsx) of the LIVE minimum-account-size analysis --
2026-08-28. See reports/_gather_live_capital_analysis.py's docstring
for the full context (a live GBPPLN position whose unrealized P&L
looked cost-negative prompted this).

Three sheets:
  - "Summary"                        -- headline numbers, capital-level
    coverage counts, at-a-glance verdict on open positions.
  - "34-Cell Minimum Account Size"   -- 17 HIGH_VOLUME_SYMBOLS pairs x
    rsi/bb, sorted by the real minimum EUR-equivalent equity needed for
    BOTH the risk gate (naturally clears 1,000 units at LIVE's current
    RISK_PCT -- 0.75% as of 2026-08-28, read live from forex.runner.
    LIVE_RISK_PCT_OVERRIDE, not hardcoded) and the cost gate (target
    profit >= 3x real round-trip cost) to pass.
  - "Open LIVE Positions"            -- every currently-open position on
    either real-money account, including legacy positions outside the
    current 17-pair universe, with real cost economics: what it needed
    to clear at entry, and whether it's cost-positive right now.

Static, computed-once values (not live formulas) -- this is a snapshot
analysis re-run on demand, not a nightly-refreshed tracker like
reports/pair_group_performance_tracker.py.

Run under py -3.12 (has openpyxl, not torch):
    python reports/_gather_live_capital_analysis.py   # phase 1
    py -3.12 reports/live_capital_analysis_report.py  # phase 2
"""
import sys, os, json
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

_cache_path = os.path.join(BASE_DIR, ".devtools", "_live_capital_analysis.json")
if not os.path.exists(_cache_path):
    print(f"ERROR: {_cache_path} not found.")
    print("Run phase 1 first: python reports/_gather_live_capital_analysis.py")
    sys.exit(1)
with open(_cache_path) as f:
    data = json.load(f)
cells, positions = data["cells"], data["positions"]
risk_pct = data.get("risk_pct", 0.0025)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill("solid", fgColor="FCE4E4")
GREEN_FILL = PatternFill("solid", fgColor="E4F7E4")
YELLOW_FILL = PatternFill("solid", fgColor="FFF6D9")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, headers, row=1):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDER


wb = openpyxl.Workbook()

# ============================================================
# Sheet 2: 34-Cell Minimum Account Size (built first so Summary can reference it)
# ============================================================
ws_c = wb.active
ws_c.title = "34-Cell Minimum Account Size"
ws_c["A1"] = f"ATOS Forex LIVE -- 34-Cell Minimum Account Size Analysis -- {NOW} PKT"
ws_c["A1"].font = Font(bold=True, size=13)
ws_c.merge_cells("A1:I1")
ws_c["A2"] = (f"At {risk_pct*100:.2f}% risk (LIVE's current RISK_PCT), never forcing a trade below 1,000 units, and requiring "
              "target profit >= 3x real round-trip cost (MIN_EDGE_TO_COST_RATIO). "
              "\"Min EUR (both gates)\" is the real minimum account equity, EUR-equivalent, needed "
              "for this pair+strategy to ever be tradeable under the current LIVE rules.")
ws_c["A2"].font = Font(italic=True, color="666666")
ws_c.merge_cells("A2:I2")
ws_c["A2"].alignment = Alignment(wrap_text=True)
headers = ["Pair", "Strategy", "ATR", "Stop Distance", "Target Distance",
           "Round-Trip Cost (EUR)", "Min Qty (cost gate)", "Min EUR (risk gate only)", "Min EUR (BOTH gates)"]
style_header(ws_c, headers, row=4)
sorted_cells = sorted(cells, key=lambda c: c["eq_eur_both_gates"])
row = 5
for c in sorted_cells:
    vals = [c["symbol"], c["strategy"].upper(), round(c["atr"], 5), round(c["stop_distance"], 5),
            round(c["target_distance"], 5), round(c["cost_eur"], 2), c["qty_cost_min"],
            round(c["eq_eur_risk_only"], 0), round(c["eq_eur_both_gates"], 0)]
    for i, v in enumerate(vals, 1):
        cell = ws_c.cell(row=row, column=i, value=v)
        cell.border = BORDER
        if i == 9:
            cell.font = BOLD
    row += 1
LAST_CELL_ROW = row - 1
ws_c.conditional_formatting.add(f"I5:I{LAST_CELL_ROW}",
    CellIsRule(operator="lessThan", formula=[str(sorted(c["eq_eur_both_gates"] for c in cells)[len(cells)//3])], fill=GREEN_FILL))
for i, w in enumerate([10, 9, 11, 13, 13, 16, 16, 18, 16], 1):
    ws_c.column_dimensions[get_column_letter(i)].width = w
ws_c.freeze_panes = "A5"
ws_c.auto_filter.ref = f"A4:I{LAST_CELL_ROW}"

# ============================================================
# Sheet 3: Open LIVE Positions -- Cost Economics
# ============================================================
ws_p = wb.create_sheet("Open LIVE Positions")
ws_p["A1"] = f"ATOS Forex LIVE -- Open Positions, Real Cost Economics -- {NOW} PKT"
ws_p["A1"].font = Font(bold=True, size=13)
ws_p.merge_cells("A1:N1")
ws_p["A2"] = ("Every currently-open position on either real-money account, including legacy "
              "positions outside the current 17-pair HIGH_VOLUME universe (from before the "
              "2026-08-28 redesign). \"Entry cost-gate ratio\" is what the trade needed to clear "
              "(>=3.0) when it opened -- \"Net if closed now\" is real unrealized P&L minus the "
              "real round-trip cost, i.e. what you'd actually keep if it closed at this instant.")
ws_p["A2"].font = Font(italic=True, color="666666")
ws_p.merge_cells("A2:N2")
ws_p["A2"].alignment = Alignment(wrap_text=True)
headers_p = ["Account", "Strategy", "Pair", "In 17-Pair Universe?", "Direction", "Qty",
             "Entry", "Current", "TP Target", "Stop", "Entry Date",
             "Entry Cost-Gate Ratio", "Unrealized P&L (EUR)", "Net If Closed Now (EUR)"]
style_header(ws_p, headers_p, row=4)
row = 5
for p in sorted(positions, key=lambda p: (p["net_if_closed_now_eur"] if p["net_if_closed_now_eur"] is not None else 0)):
    in_universe = "Yes" if p["is_high_volume"] else "No (legacy)"
    vals = [p["account"], p["strategy"].upper(), p["symbol"], in_universe, p["direction"], p["quantity"],
            p["entry_price"], p["current_price"], p["tp_price"], p["stop_price"], p["entry_date"],
            round(p["entry_cost_gate_ratio"], 2) if p["entry_cost_gate_ratio"] is not None else "N/A",
            round(p["unrealized_eur"], 2) if p["unrealized_eur"] is not None else "N/A",
            round(p["net_if_closed_now_eur"], 2) if p["net_if_closed_now_eur"] is not None else "N/A"]
    for i, v in enumerate(vals, 1):
        cell = ws_p.cell(row=row, column=i, value=v)
        cell.border = BORDER
        if i == 4 and v.startswith("No"):
            cell.font = Font(italic=True, color="B45309")
        if i == 14 and isinstance(v, (int, float)):
            cell.fill = GREEN_FILL if v >= 0 else RED_FILL
            cell.font = BOLD
    row += 1
LAST_POS_ROW = row - 1
for i, w in enumerate([14, 9, 8, 16, 10, 8, 11, 11, 11, 11, 12, 16, 16, 18], 1):
    ws_p.column_dimensions[get_column_letter(i)].width = w
ws_p.freeze_panes = "A5"

# ============================================================
# Sheet 1: Summary (inserted first)
# ============================================================
ws_s = wb.create_sheet("Summary", 0)
ws_s["A1"] = f"ATOS Forex LIVE -- Capital Requirements Summary -- {NOW} PKT"
ws_s["A1"].font = Font(bold=True, size=14)
ws_s.merge_cells("A1:D1")

cheapest = sorted_cells[0]
most_expensive = sorted_cells[-1]
CURRENT_CAP_EUR = 1350  # forex_live_eur's configured cap, config/capital.json -- marked in the table below
thresholds = sorted(set([500, 900, 1000, CURRENT_CAP_EUR, 1500, 2000, 3000, 4000, 5000, 6500]))
coverage = {t: sum(1 for c in cells if c["eq_eur_both_gates"] <= t) for t in thresholds}

r0 = 3
ws_s.cell(row=r0, column=1, value="Cheapest tradeable cell (both gates)").font = BOLD
ws_s.cell(row=r0, column=2, value=f"{cheapest['symbol']} / {cheapest['strategy'].upper()} -- EUR {cheapest['eq_eur_both_gates']:,.0f}")
r0 += 1
ws_s.cell(row=r0, column=1, value="Most expensive cell (both gates)").font = BOLD
ws_s.cell(row=r0, column=2, value=f"{most_expensive['symbol']} / {most_expensive['strategy'].upper()} -- EUR {most_expensive['eq_eur_both_gates']:,.0f}")
r0 += 1
ws_s.cell(row=r0, column=1, value=f"(all figures below at {risk_pct*100:.2f}% risk -- LIVE's current RISK_PCT)").font = Font(italic=True, color="666666")
r0 += 1
ws_s.cell(row=r0, column=1, value="Capital level").font = BOLD
ws_s.cell(row=r0, column=2, value="Cells tradeable / 34").font = BOLD
r0 += 1
for t in thresholds:
    label = f"EUR {t:,}" + ("  <- current configured cap" if t == CURRENT_CAP_EUR else "")
    ws_s.cell(row=r0, column=1, value=label)
    if t == CURRENT_CAP_EUR:
        ws_s.cell(row=r0, column=1).font = BOLD
    ws_s.cell(row=r0, column=2, value=f"{coverage[t]} / 34")
    if coverage[t] == 0:
        ws_s.cell(row=r0, column=2).font = Font(color="B91C1C")
    elif coverage[t] == 34:
        ws_s.cell(row=r0, column=2).font = Font(color="15803D", bold=True)
    r0 += 1

r0 += 1
ws_s.cell(row=r0, column=1, value="Open LIVE positions right now").font = BOLD
r0 += 1
net_negative = [p for p in positions if p["net_if_closed_now_eur"] is not None and p["net_if_closed_now_eur"] < 0]
net_positive = [p for p in positions if p["net_if_closed_now_eur"] is not None and p["net_if_closed_now_eur"] >= 0]
ws_s.cell(row=r0, column=1, value="Total open positions")
ws_s.cell(row=r0, column=2, value=len(positions))
r0 += 1
ws_s.cell(row=r0, column=1, value="Would be net-negative if closed right now")
ws_s.cell(row=r0, column=2, value=f"{len(net_negative)} ({', '.join(p['symbol'] for p in net_negative)})" if net_negative else "0")
if net_negative:
    ws_s.cell(row=r0, column=2).font = Font(color="B91C1C")
r0 += 1
ws_s.cell(row=r0, column=1, value="Would be net-positive if closed right now")
ws_s.cell(row=r0, column=2, value=f"{len(net_positive)} ({', '.join(p['symbol'] for p in net_positive)})" if net_positive else "0")
if net_positive:
    ws_s.cell(row=r0, column=2).font = Font(color="15803D")
r0 += 1
outside = [p for p in positions if not p["is_high_volume"]]
ws_s.cell(row=r0, column=1, value="Positions outside the current 17-pair universe (legacy)")
ws_s.cell(row=r0, column=2, value=f"{len(outside)} ({', '.join(p['symbol'] for p in outside)})" if outside else "0")

r0 += 2
ws_s.cell(row=r0, column=1, value="Note").font = BOLD
ws_s.cell(row=r0, column=2, value=("\"Net if closed now\" being negative does NOT mean the cost gate failed -- it "
                                    "means the trade hasn't reached its profit target yet. Every open position's "
                                    "entry cost-gate ratio (see 'Open LIVE Positions' sheet) was checked against the "
                                    "3.0x minimum at entry, using the target price, not the current price."))
ws_s.cell(row=r0, column=2).alignment = Alignment(wrap_text=True)
ws_s.merge_cells(f"B{r0}:D{r0}")

ws_s.column_dimensions["A"].width = 46
ws_s.column_dimensions["B"].width = 46
ws_s.column_dimensions["C"].width = 20
ws_s.column_dimensions["D"].width = 20

out_path = os.path.join(DATA_DIR, "forex_live_capital_analysis.xlsx")
wb.save(out_path)
print("Saved:", out_path)
