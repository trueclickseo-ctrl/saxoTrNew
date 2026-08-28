"""
reports/module_performance_tracker.py <module> <display-name> <currency>
---------------------------------------------------------------------------
Generalized daily PF/P&L/WR performance tracker for any pnl_ledger.db
module OTHER than forex (which has its own, richer
reports/pair_group_performance_tracker.py -- forex needs live re-priced
EUR conversion across many quote currencies + a Forex Grouping tier
breakdown that doesn't apply anywhere else). Built 2026-08-28, explicit
user request: "also make advance Excel tracker for ETF, Stocks, Futures
-- keep track daily, weekly, monthly like we have advance Excel tracker
for Forex."

Unlike forex's tracker, this needs NO live-repricing phase-1 step --
every module here (futures/etf/stock) is single-currency and
pnl_ledger.db's `trades` table already stores the real dealt
realized_pnl (net, post-commission) and commission per closed trade, so
this is a single self-contained script (sqlite3 + openpyxl only, no
torch/forex.runner import at all).

Usage:
    py -3.12 reports/module_performance_tracker.py futures Futures USD
    py -3.12 reports/module_performance_tracker.py etf     ETF     USD
    py -3.12 reports/module_performance_tracker.py stock   Stocks  SEK

Builds/overwrites ONE persistent workbook per module:
    data/{module}_performance_tracker.xlsx
with 4 sheets: Per-Strategy Performance, Per-Symbol Performance,
Daily/Weekly/Monthly Performance (one sheet each, added below the first
two) -- same live-formula-over-a-hidden-Trade-Detail-sheet design as
forex's tracker, see reports/_perf_common.py.

Read-only analysis only -- never touches a live signal, gate, stop, or
order. Safe to run any time.
"""
import sys, os, sqlite3
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
from _perf_common import (parse_close_date, day_key, week_key, month_key,
                           sorted_distinct_period_keys, write_metric_formulas)

if len(sys.argv) < 4:
    print("Usage: py -3.12 reports/module_performance_tracker.py <module> <DisplayName> <CURRENCY>")
    print("  e.g.: py -3.12 reports/module_performance_tracker.py futures Futures USD")
    sys.exit(1)
MODULE, DISPLAY, CCY = sys.argv[1], sys.argv[2], sys.argv[3]

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill("solid", fgColor="FCE4E4")
GREEN_FILL = PatternFill("solid", fgColor="E4F7E4")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, headers, row=1):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDER


# ============================================================
# Pull directly from pnl_ledger.db -- no live re-pricing needed, every
# closed row already has its real dealt realized_pnl (net) + commission.
# ============================================================
conn = sqlite3.connect(os.path.join(DATA_DIR, "pnl_ledger.db"))
conn.row_factory = sqlite3.Row
rows = conn.execute(
    "SELECT * FROM trades WHERE module=? ORDER BY id", (MODULE,)
).fetchall()
conn.close()

usable, unresolved = [], []
for row in rows:
    net = row["realized_pnl"]
    comm = row["commission"] or 0.0
    if row["status"] == "closed" and net is None:
        # A closed trade whose real P&L is genuinely unknown (e.g.
        # futures id 60/CL -- broker-side ambiguity, no confirmed closing
        # fill, documented in pnl_tracker's own test as a deliberately
        # honest non-fabricated record, never to be guessed at). Track
        # separately so it's still SHOWN somewhere (as "P&L UNKNOWN"), not
        # silently dropped -- silently vanishing from every sheet would
        # look like missing data instead of a documented open question.
        unresolved.append({"strategy": row["strategy"] or "—", "symbol": row["symbol"]})
        continue
    if row["status"] != "closed":
        # Open positions have no realized P&L yet -- they don't belong on
        # a PF/WR/Net-P&L performance sheet (unrealized P&L needs a live
        # price, which is exactly the complexity this script deliberately
        # avoids; the module's own live dashboard already shows open P&L).
        continue
    usable.append({
        "strategy": row["strategy"] or "—",
        "symbol": row["symbol"],
        "net_pnl": round(net, 2),
        "commission": round(comm, 2),
        "gross_pnl": round(net + comm, 2),
        "timestamp_close": row["timestamp_close"],
    })
print(f"  {len(usable)}/{len(rows)} trades usable (closed only, {len(unresolved)} closed-but-unknown-P&L excluded)")

ALL_STRATEGIES = sorted({t["strategy"] for t in usable} | {u["strategy"] for u in unresolved})
ALL_SYMBOLS = sorted({t["symbol"] for t in usable} | {u["symbol"] for u in unresolved})
UNRESOLVED_SYMBOLS = {u["symbol"] for u in unresolved} - {t["symbol"] for t in usable}
UNRESOLVED_STRATEGIES = {u["strategy"] for u in unresolved} - {t["strategy"] for t in usable}

wb = openpyxl.Workbook()

# ============================================================
# Hidden "Trade Detail" sheet -- same column layout as forex's tracker
# (Group column left blank -- this module has no tier system).
# ============================================================
ws = wb.active
ws.title = "Trade Detail"
style_header(ws, ["Strategy", "Symbol", "Group", "Direction", "Units", "Status",
                   f"Gross P&L ({CCY})", f"Commission ({CCY})", f"Net P&L ({CCY})", "Net Result",
                   "Close Date", "Week", "Month"])
row = 2
for t in usable:
    net_result = "WIN" if t["net_pnl"] > 0 else "LOSS"
    d = parse_close_date(t["timestamp_close"])
    vals = [t["strategy"], t["symbol"], "", "", "", "closed",
            t["gross_pnl"], t["commission"], t["net_pnl"], net_result,
            day_key(d) if d else "", week_key(d) if d else "", month_key(d) if d else ""]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BORDER
        if i == 10:
            c.fill = GREEN_FILL if net_result == "WIN" else RED_FILL
    row += 1
LAST_ROW = row - 1
for i, w in enumerate([14, 10, 8, 10, 8, 8, 14, 14, 12, 10, 12, 10, 10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.sheet_state = "hidden"
TD = "'Trade Detail'"

HEADERS = ["Strategy/Symbol", "Trades", "Wins (net)", "WR %",
           f"Gross P&L ({CCY})", f"Commission ({CCY})", f"Net P&L ({CCY})", "Profit Factor"]
COL_WIDTHS = [22, 9, 11, 8, 15, 15, 14, 13]


def build_key_sheet(title, td_col_letter, keys, index=None, unresolved_keys=frozenset()):
    ws2 = wb.create_sheet(title, index) if index is not None else wb.create_sheet(title)
    ws2["A1"] = f"ATOS {DISPLAY} -- {title} -- last updated {NOW} PKT ({CCY})"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.merge_cells("A1:H1")
    style_header(ws2, HEADERS, row=3)
    r = 4
    for key in keys:
        ws2.cell(row=r, column=1, value=key).font = BOLD
        if key in unresolved_keys:
            # A closed trade exists but its real P&L was never confirmed
            # (see the module docstring/comment above `unresolved`) --
            # show that honestly instead of a live-formula "0 trades",
            # which would look like no activity happened at all.
            ws2.cell(row=r, column=2, value="P&L UNKNOWN (see docs)").font = Font(italic=True, color="999999")
            for c in range(3, 9):
                ws2.cell(row=r, column=c, value="—").font = Font(italic=True, color="999999")
        else:
            crit = f'{TD}!${td_col_letter}$2:${td_col_letter}${LAST_ROW},"{key}"'
            write_metric_formulas(ws2, r, crit, TD, LAST_ROW, start_col=2)
        for c in range(1, 9):
            ws2.cell(row=r, column=c).border = BORDER
        r += 1
    last = r - 1
    if last >= 4:
        ws2.conditional_formatting.add(f"G4:G{last}", CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL))
        ws2.conditional_formatting.add(f"G4:G{last}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL))
    else:
        ws2.cell(row=4, column=1, value="No closed trades yet.").font = Font(italic=True, color="999999")
    for i, w in enumerate(COL_WIDTHS, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A4"
    return ws2


# ============================================================
# Sheet 1: Per-Strategy Performance
# ============================================================
build_key_sheet("Per-Strategy Performance", "A", ALL_STRATEGIES, index=0,
                 unresolved_keys=UNRESOLVED_STRATEGIES)

# ============================================================
# Sheet 2: Per-Symbol Performance
# ============================================================
build_key_sheet("Per-Symbol Performance", "B", ALL_SYMBOLS,
                 unresolved_keys=UNRESOLVED_SYMBOLS)

# ============================================================
# Sheets 3-5: Daily / Weekly / Monthly Performance
# ============================================================
build_key_sheet("Daily Performance", "K", sorted_distinct_period_keys(usable, day_key))
build_key_sheet("Weekly Performance", "L", sorted_distinct_period_keys(usable, week_key))
build_key_sheet("Monthly Performance", "M", sorted_distinct_period_keys(usable, month_key))

out_path = os.path.join(DATA_DIR, f"{MODULE}_performance_tracker.xlsx")
wb.save(out_path)
print(f"Saved (single persistent file, overwritten in place): {out_path}")
