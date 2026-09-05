"""
reports/pair_group_performance_tracker.py
-------------------------------------------
Daily PF / P&L / Win-Rate performance tracker -- ONE persistent workbook,
overwritten in place every run (not a new dated file each day), per
explicit user request 2026-08-28: "save data/performance PF and PNL and
WR for the new pairs and all Existing pairs, update in a excel sheet,
keep that sheet with you... Track the performance of Strategies Group
Wise and Track the Performance of Pairs wise... Important do this
everyday when the trading is close."

Five sheets, all built from the same trade-level data as
reports/daily_sim_report.py:
  - "Per-Group Performance"  -- one row per Forex Grouping tier (the
    EXACT names forex_dashboard.py uses: High Volume, Core Standard,
    Scandi, Metals, Exotic Asia, Exotic Europe, Exotic High-Yield/Carry,
    Exotic Latam/Mideast), so this is directly reusable when configuring
    ATOS LIVE later.
  - "Per-Pair Performance"   -- one row per individual pair (all 184,
    including pairs with zero trades so far -- explicitly listed as
    "NO DATA" rather than silently absent, so a newly-added pair's
    coverage gap is visible, not invisible).
  - "Daily/Weekly/Monthly Performance" (added 2026-08-28, explicit user
    request -- "Important do this everyday when the trading is close...
    keep track daily, weekly, monthly") -- one row per calendar day / ISO
    week / month that has at least one CLOSED trade, oldest first. Only
    closed trades are bucketed (an open position has no "when this
    happened" close date yet) -- see reports/_perf_common.py.

Same two-phase split as daily_sim_report.py/live_readiness_report.py
(forex.runner transitively imports torch, not installed on whichever
Python has openpyxl here):

    python reports/_gather_daily_sim_data.py          # phase 1: real Saxo data
    py -3.12 reports/pair_group_performance_tracker.py # phase 2: build/update the .xlsx

Read-only analysis only -- never touches a live signal, gate, stop, or
order. Safe to run any time, including during the forward-SIM
"no-touch" observation period (see forex_live_pnl_base_currency_bug_
2026-08-26.md -- architecture frozen, this doesn't change any of it).
"""
import sys, os, json
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
from _perf_common import (parse_close_date, day_key, week_key, month_key,
                           sorted_distinct_period_keys, write_metric_formulas, METRIC_HEADERS)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

_trades_cache_path = os.path.join(BASE_DIR, ".devtools", "_daily_sim_trades.json")
if not os.path.exists(_trades_cache_path):
    print(f"ERROR: {_trades_cache_path} not found.")
    print("Run phase 1 first:  python reports/_gather_daily_sim_data.py")
    sys.exit(1)
with open(_trades_cache_path) as f:
    trades = json.load(f)

# Full pair/group universe (so pairs with zero trades still show a row,
# not a silent gap) -- read via a plain JSON/py-level import that doesn't
# need forex.runner/torch, matching this file's own no-torch Python.
sys.path.insert(0, BASE_DIR)  # forex/universe.py has no torch dependency
from forex.universe import PAIRS, get_group  # noqa: E402

ALL_PAIRS_BY_GROUP = {}
for p in PAIRS:
    ALL_PAIRS_BY_GROUP.setdefault(get_group(p["symbol"]), []).append(p["symbol"])

# Canonical Forex Grouping order (matches forex_dashboard.py's section order)
GROUP_ORDER = ["High Volume", "Core Standard", "Scandi", "Metals",
               "Exotic Asia", "Exotic Europe", "Exotic High-Yield/Carry", "Exotic Latam/Mideast"]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill("solid", fgColor="FCE4E4")
GREEN_FILL = PatternFill("solid", fgColor="E4F7E4")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, headers, row=1):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDER


usable = [t for t in trades if t["commission_eur"] is not None and t["net_pnl_eur"] is not None]
closed = [t for t in usable if t["status"] == "closed"]
print(f"  {len(usable)}/{len(trades)} trades usable ({len(closed)} closed)")

wb = openpyxl.Workbook()

# ============================================================
# Hidden "Trade Detail" sheet -- raw data the two summary sheets' live
# formulas (COUNTIF/SUMIF/SUMIFS) read from. Kept for full traceability
# (any PF/WR number on the summary sheets can be traced back to its
# underlying trades) without cluttering the two sheets the user actually
# looks at day to day.
ws = wb.active
ws.title = "Trade Detail"
style_header(ws, ["Strategy", "Pair", "Group", "Direction", "Units", "Status",
                   "Gross P&L (EUR)", "Commission (EUR)", "Net P&L (EUR)", "Net Result",
                   "Close Date", "Week", "Month"])
row = 2
for t in usable:
    net = t["net_pnl_eur"]
    net_result = "WIN" if net > 0 else "LOSS"
    d = parse_close_date(t.get("timestamp_close"))
    vals = [t["strategy"], t["symbol"], t["group"], t["direction"], t["quantity"], t["status"],
            t["gross_pnl_eur"], t["commission_eur"], net, net_result,
            day_key(d) if d else "", week_key(d) if d else "", month_key(d) if d else ""]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BORDER
        if i == 10:
            c.fill = GREEN_FILL if net_result == "WIN" else RED_FILL
    row += 1
LAST_ROW = row - 1
for i, w in enumerate([12, 9, 24, 10, 10, 8, 14, 14, 12, 10, 12, 10, 10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.sheet_state = "hidden"
TD = "'Trade Detail'"


def _perf_formulas(ws2, row_i, group_col_letter, key_cell):
    """Trades / Wins / WR% / Gross / Commission / Net / PF, all live
    formulas reading from Trade Detail so this sheet self-updates if
    Trade Detail is ever regenerated without rebuilding this sheet."""
    crit = f"{TD}!${group_col_letter}$2:${group_col_letter}${LAST_ROW},{key_cell}"
    ws2.cell(row=row_i, column=2, value=f'=COUNTIF({crit})')
    ws2.cell(row=row_i, column=3, value=f'=COUNTIFS({crit},{TD}!$J$2:$J${LAST_ROW},"WIN")')
    ws2.cell(row=row_i, column=4, value=f'=IFERROR(ROUND(C{row_i}/B{row_i}*100,1),"")')
    ws2.cell(row=row_i, column=5, value=f'=ROUND(SUMIF({crit},{TD}!$G$2:$G${LAST_ROW}),2)')
    ws2.cell(row=row_i, column=6, value=f'=ROUND(SUMIF({crit},{TD}!$H$2:$H${LAST_ROW}),2)')
    ws2.cell(row=row_i, column=7, value=f'=ROUND(SUMIF({crit},{TD}!$I$2:$I${LAST_ROW}),2)')
    # Profit Factor = gross profit of winning trades / abs(gross loss of losing trades),
    # using NET P&L (post-commission) so PF reflects real tradeable edge, not gross.
    win_sum = f'SUMIFS({TD}!$I$2:$I${LAST_ROW},{crit},{TD}!$I$2:$I${LAST_ROW},">0")'
    loss_sum = f'SUMIFS({TD}!$I$2:$I${LAST_ROW},{crit},{TD}!$I$2:$I${LAST_ROW},"<0")'
    ws2.cell(row=row_i, column=8, value=f'=IFERROR(ROUND({win_sum}/ABS({loss_sum}),2),IF(B{row_i}=0,"",">0 (no losers)"))')
    for c in range(1, 9):
        ws2.cell(row=row_i, column=c).border = BORDER


HEADERS = ["Group/Pair", "Trades", "Wins (net)", "WR %", "Gross P&L (EUR)",
           "Commission (EUR)", "Net P&L (EUR)", "Profit Factor"]
COL_WIDTHS = [26, 9, 11, 8, 15, 15, 14, 13]

# ============================================================
# Sheet 1: Per-Group Performance (the EXACT Forex Grouping tier names)
# ============================================================
ws_g = wb.create_sheet("Per-Group Performance", 0)
ws_g["A1"] = f"ATOS Forex -- Performance by Group -- last updated {NOW} PKT"
ws_g["A1"].font = Font(bold=True, size=13)
ws_g.merge_cells("A1:H1")
style_header(ws_g, HEADERS, row=3)
r = 4
for group in GROUP_ORDER:
    n_pairs = len(ALL_PAIRS_BY_GROUP.get(group, []))
    ws_g.cell(row=r, column=1, value=f"{group} ({n_pairs} pairs)").font = BOLD
    _perf_formulas(ws_g, r, "C", f'"{group}"')
    r += 1
LAST_GROUP_ROW = r - 1
ws_g.conditional_formatting.add(f"G4:G{LAST_GROUP_ROW}", CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL))
ws_g.conditional_formatting.add(f"G4:G{LAST_GROUP_ROW}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL))
for i, w in enumerate(COL_WIDTHS, 1):
    ws_g.column_dimensions[get_column_letter(i)].width = w
ws_g.freeze_panes = "A4"

# ============================================================
# Sheet 2: Per-Pair Performance (every pair in the universe, incl. zero-trade ones)
# ============================================================
ws_p = wb.create_sheet("Per-Pair Performance")
ws_p["A1"] = f"ATOS Forex -- Performance by Pair -- last updated {NOW} PKT -- {len(PAIRS)} total pairs"
ws_p["A1"].font = Font(bold=True, size=13)
ws_p.merge_cells("A1:H1")
style_header(ws_p, ["Pair", "Group"] + HEADERS[1:], row=3)
r = 4
traded_symbols = {t["symbol"] for t in usable}
for group in GROUP_ORDER:
    for sym in sorted(ALL_PAIRS_BY_GROUP.get(group, [])):
        ws_p.cell(row=r, column=1, value=sym).font = BOLD
        ws_p.cell(row=r, column=2, value=group)
        if sym in traded_symbols:
            # Per-pair sheet has one extra leading "Group" column vs the
            # per-group sheet, so _perf_formulas' 8-column layout (Trades..PF)
            # is written shifted right by one -- pass column letters B (pair
            # symbol, matches Trade Detail!B) and a wrapper that writes into
            # columns 3-9 instead of 2-8.
            crit = f'{TD}!$B$2:$B${LAST_ROW},A{r}'
            ws_p.cell(row=r, column=3, value=f'=COUNTIF({crit})')
            ws_p.cell(row=r, column=4, value=f'=COUNTIFS({crit},{TD}!$J$2:$J${LAST_ROW},"WIN")')
            ws_p.cell(row=r, column=5, value=f'=IFERROR(ROUND(D{r}/C{r}*100,1),"")')
            ws_p.cell(row=r, column=6, value=f'=ROUND(SUMIF({crit},{TD}!$G$2:$G${LAST_ROW}),2)')
            ws_p.cell(row=r, column=7, value=f'=ROUND(SUMIF({crit},{TD}!$H$2:$H${LAST_ROW}),2)')
            ws_p.cell(row=r, column=8, value=f'=ROUND(SUMIF({crit},{TD}!$I$2:$I${LAST_ROW}),2)')
            win_sum = f'SUMIFS({TD}!$I$2:$I${LAST_ROW},{crit},{TD}!$I$2:$I${LAST_ROW},">0")'
            loss_sum = f'SUMIFS({TD}!$I$2:$I${LAST_ROW},{crit},{TD}!$I$2:$I${LAST_ROW},"<0")'
            ws_p.cell(row=r, column=9, value=f'=IFERROR(ROUND({win_sum}/ABS({loss_sum}),2),IF(C{r}=0,"",">0 (no losers)"))')
        else:
            ws_p.cell(row=r, column=3, value=0)
            for c in range(4, 9):
                ws_p.cell(row=r, column=c, value="NO DATA").font = Font(italic=True, color="999999")
        for c in range(1, 10):
            ws_p.cell(row=r, column=c).border = BORDER
        r += 1
LAST_PAIR_ROW = r - 1
ws_p.conditional_formatting.add(f"H4:H{LAST_PAIR_ROW}", CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL))
ws_p.conditional_formatting.add(f"H4:H{LAST_PAIR_ROW}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL))
for i, w in enumerate([10, 24, 9, 11, 8, 15, 15, 14, 13], 1):
    ws_p.column_dimensions[get_column_letter(i)].width = w
ws_p.freeze_panes = "A4"
ws_p.auto_filter.ref = f"A3:I{LAST_PAIR_ROW}"

# ============================================================
# Sheets 3-5: Daily / Weekly / Monthly Performance (2026-08-28)
# ============================================================
TIME_HEADERS = ["Period"] + METRIC_HEADERS
TIME_COL_WIDTHS = [14, 9, 11, 8, 15, 15, 14, 13]


def build_time_sheet(title, td_col_letter, key_fn):
    keys = sorted_distinct_period_keys(usable, key_fn)
    ws_t = wb.create_sheet(title)
    ws_t["A1"] = f"ATOS Forex -- {title} -- last updated {NOW} PKT"
    ws_t["A1"].font = Font(bold=True, size=13)
    ws_t.merge_cells("A1:H1")
    style_header(ws_t, TIME_HEADERS, row=3)
    if not keys:
        ws_t.cell(row=4, column=1, value="No closed trades with a known close date yet.").font = Font(italic=True, color="999999")
        for i, w in enumerate(TIME_COL_WIDTHS, 1):
            ws_t.column_dimensions[get_column_letter(i)].width = w
        return
    r = 4
    for key in keys:
        ws_t.cell(row=r, column=1, value=key).font = BOLD
        crit = f"{TD}!${td_col_letter}$2:${td_col_letter}${LAST_ROW},\"{key}\""
        write_metric_formulas(ws_t, r, crit, TD, LAST_ROW, start_col=2)
        for c in range(1, 9):
            ws_t.cell(row=r, column=c).border = BORDER
        r += 1
    last_row_t = r - 1
    ws_t.conditional_formatting.add(f"G4:G{last_row_t}", CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL))
    ws_t.conditional_formatting.add(f"G4:G{last_row_t}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL))
    for i, w in enumerate(TIME_COL_WIDTHS, 1):
        ws_t.column_dimensions[get_column_letter(i)].width = w
    ws_t.freeze_panes = "A4"


build_time_sheet("Daily Performance", "K", day_key)
build_time_sheet("Weekly Performance", "L", week_key)
build_time_sheet("Monthly Performance", "M", month_key)

# ============================================================
# Sheet 6: Per-Pair × Strategy  (only pairs with 2+ strategies)
# One row per (pair, strategy) combo so the user can see which strategy
# is carrying / dragging each multi-strategy pair.
# ============================================================
ws_ps = wb.create_sheet("Per-Pair × Strategy")
ws_ps["A1"] = (
    f"ATOS Forex -- Per-Pair × Strategy -- last updated {NOW} PKT  "
    f"(pairs with 2+ strategies only)"
)
ws_ps["A1"].font = Font(bold=True, size=13)
ws_ps.merge_cells("A1:J1")

PS_HEADERS = ["Pair", "Group", "Strategy", "Trades", "Wins",
              "WR %", "Gross P&L (EUR)", "Commission (EUR)", "Net P&L (EUR)", "Profit Factor"]
style_header(ws_ps, PS_HEADERS, row=3)
PS_COL_WIDTHS = [10, 24, 28, 9, 7, 7, 15, 15, 14, 13]
for i, w in enumerate(PS_COL_WIDTHS, 1):
    ws_ps.column_dimensions[get_column_letter(i)].width = w

# Build (pair → {strategy → group}) lookup from Trade Detail data
from collections import defaultdict
pair_strategy_group: dict = defaultdict(dict)  # pair → strategy → group
for t in usable:
    pair_strategy_group[t["symbol"]][t["strategy"]] = t.get("group", "")

# Only pairs that appear with 2+ distinct strategies in closed trades
multi_strat_pairs = sorted(
    sym for sym, strats in pair_strategy_group.items() if len(strats) >= 2
)

ps_row = 4
prev_pair = None
for sym in multi_strat_pairs:
    strategies = sorted(pair_strategy_group[sym].keys())
    group = next(iter(pair_strategy_group[sym].values()), "")
    for strat in strategies:
        # Shade alternate pairs for readability
        is_alt = (multi_strat_pairs.index(sym) % 2 == 1)
        alt_fill = PatternFill("solid", fgColor="F0F4F8") if is_alt else None

        ws_ps.cell(row=ps_row, column=1, value=sym if strat == strategies[0] else "").font = BOLD
        ws_ps.cell(row=ps_row, column=2, value=group if strat == strategies[0] else "")
        ws_ps.cell(row=ps_row, column=3, value=strat)

        # COUNTIFS / SUMIFS keyed on (Pair=col B, Strategy=col A, Net Result=col J)
        pair_range = f"{TD}!$B$2:$B${LAST_ROW}"
        strat_range = f"{TD}!$A$2:$A${LAST_ROW}"
        result_range = f"{TD}!$J$2:$J${LAST_ROW}"
        gross_range = f"{TD}!$G$2:$G${LAST_ROW}"
        comm_range = f"{TD}!$H$2:$H${LAST_ROW}"
        net_range = f"{TD}!$I$2:$I${LAST_ROW}"
        pair_crit = f'"{sym}"'
        strat_crit = f'"{strat}"'
        base_crit = f"{pair_range},{pair_crit},{strat_range},{strat_crit}"

        ws_ps.cell(row=ps_row, column=4,
                   value=f'=COUNTIFS({base_crit})')
        ws_ps.cell(row=ps_row, column=5,
                   value=f'=COUNTIFS({base_crit},{result_range},"WIN")')
        ws_ps.cell(row=ps_row, column=6,
                   value=f'=IFERROR(ROUND(E{ps_row}/D{ps_row}*100,1),"")')
        ws_ps.cell(row=ps_row, column=7,
                   value=f'=ROUND(SUMIFS({gross_range},{pair_range},{pair_crit},{strat_range},{strat_crit}),2)')
        ws_ps.cell(row=ps_row, column=8,
                   value=f'=ROUND(SUMIFS({comm_range},{pair_range},{pair_crit},{strat_range},{strat_crit}),2)')
        ws_ps.cell(row=ps_row, column=9,
                   value=f'=ROUND(SUMIFS({net_range},{pair_range},{pair_crit},{strat_range},{strat_crit}),2)')
        win_sum = (f'SUMIFS({net_range},{pair_range},{pair_crit},'
                   f'{strat_range},{strat_crit},{net_range},">0")')
        loss_sum = (f'SUMIFS({net_range},{pair_range},{pair_crit},'
                    f'{strat_range},{strat_crit},{net_range},"<0")')
        ws_ps.cell(row=ps_row, column=10,
                   value=f'=IFERROR(ROUND({win_sum}/ABS({loss_sum}),2),'
                         f'IF(D{ps_row}=0,"",">0 (no losers)"))')

        for col in range(1, 11):
            c = ws_ps.cell(row=ps_row, column=col)
            c.border = BORDER
            if alt_fill:
                c.fill = alt_fill

        ps_row += 1

    # Thin separator row between pairs (empty row with just a bottom border)
    for col in range(1, 11):
        ws_ps.cell(row=ps_row, column=col).border = Border(bottom=THIN)
    ps_row += 1

LAST_PS_ROW = ps_row - 1
# Green/red on Net P&L column (col I = column 9)
ws_ps.conditional_formatting.add(
    f"I4:I{LAST_PS_ROW}",
    CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL)
)
ws_ps.conditional_formatting.add(
    f"I4:I{LAST_PS_ROW}",
    CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL)
)
ws_ps.freeze_panes = "A4"
ws_ps.auto_filter.ref = f"A3:J3"

out_path = os.path.join(DATA_DIR, "forex_performance_tracker.xlsx")
wb.save(out_path)
print("Saved (single persistent file, overwritten in place):", out_path)
