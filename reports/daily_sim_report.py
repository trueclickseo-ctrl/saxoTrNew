"""
reports/daily_sim_report.py
----------------------------
Daily end-of-day SIM analysis workbook. Run this any day to get a fresh,
date-stamped snapshot in Downloads: SIM trade/cost/strategy analysis
PLUS the three forward-SIM observation logs added 2026-08-27
(forex/forward_observation.py): cost-gate decisions, currency exposure
snapshots, and entry/exit trade cards.

Two-phase design -- forex.runner transitively imports torch (for
strategy_cnn_lstm), which isn't installed on whichever Python has
openpyxl in this environment, so this can't be one process:

    python reports/_gather_daily_sim_data.py   # phase 1: real Saxo data
    py -3.12 reports/daily_sim_report.py        # phase 2: build the .xlsx

Everything here is read-only analysis -- it never touches a live signal,
gate, stop, or order. Safe to run any time, including during the
forward-SIM "no-touch" observation period.
"""
import sys, os, json
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
TODAY = datetime.now().strftime("%Y-%m-%d")

_trades_cache_path = os.path.join(BASE_DIR, ".devtools", "_daily_sim_trades.json")
if not os.path.exists(_trades_cache_path):
    print(f"ERROR: {_trades_cache_path} not found.")
    print("Run phase 1 first:  python reports/_gather_daily_sim_data.py")
    sys.exit(1)
with open(_trades_cache_path) as f:
    trades = json.load(f)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill("solid", fgColor="FCE4E4")
GREEN_FILL = PatternFill("solid", fgColor="E4F7E4")
YELLOW_FILL = PatternFill("solid", fgColor="FFF6D9")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, headers, row=1):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDER


def load_jsonl(path):
    if not os.path.exists(path):
        return []
    rows = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    rows.append(json.loads(line))
                except Exception:
                    pass
    return rows


# ============================================================
# 1. Trades already gathered by phase 1 (_gather_daily_sim_data.py)
# ============================================================
usable = [t for t in trades if t["commission_eur"] is not None and t["gross_pnl_eur"] is not None]
print(f"  {len(usable)}/{len(trades)} trades usable")

# ============================================================
# 2. Build the workbook
# ============================================================
wb = openpyxl.Workbook()

ws = wb.active
ws.title = "Trade Detail"
style_header(ws, ["Strategy", "Pair", "Direction", "Units", "Entry", "Exit/Current", "Status",
                   "Exit Reason", "Gross P&L (EUR)", "Commission (EUR)", "Net P&L (EUR)",
                   "Gross Result", "Net Result", "Flipped?"])
row = 2
for t in usable:
    gross, comm, net = t["gross_pnl_eur"], t["commission_eur"], t["net_pnl_eur"]
    gross_result = "WIN" if gross > 0 else "LOSS"
    net_result = "WIN" if net > 0 else "LOSS"
    flipped = "YES" if gross_result != net_result else "no"
    vals = [t["strategy"], t["symbol"], t["direction"], t["quantity"], t["entry"], t["exit"],
            t["status"], t.get("exit_reason", ""), gross, comm, net, gross_result, net_result, flipped]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BORDER
        if i == 13:
            c.fill = GREEN_FILL if net_result == "WIN" else RED_FILL
        if i == 14 and flipped == "YES":
            c.fill, c.font = YELLOW_FILL, BOLD
    row += 1
LAST_ROW = row - 1
for i, w in enumerate([12, 9, 10, 9, 11, 12, 8, 24, 13, 13, 13, 11, 10, 9], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
TD = "'Trade Detail'"

# --- Per-Pair / Per-Strategy summaries, live formulas ---
def build_summary_sheet(title, group_col_letter, group_values):
    ws2 = wb.create_sheet(title)
    style_header(ws2, [title.split()[1], "Trades", "Gross Wins", "Net Wins", "Flipped",
                        "Total Gross (EUR)", "Total Commission (EUR)", "Total Net (EUR)",
                        "Commission % |Gross|", "Verdict"])
    for i, val in enumerate(sorted(group_values), 2):
        ws2.cell(row=i, column=1, value=val)
        ws2.cell(row=i, column=2, value=f'=COUNTIF({TD}!${group_col_letter}$2:${group_col_letter}${LAST_ROW},A{i})')
        ws2.cell(row=i, column=3, value=f'=COUNTIFS({TD}!${group_col_letter}$2:${group_col_letter}${LAST_ROW},A{i},{TD}!$L$2:$L${LAST_ROW},"WIN")')
        ws2.cell(row=i, column=4, value=f'=COUNTIFS({TD}!${group_col_letter}$2:${group_col_letter}${LAST_ROW},A{i},{TD}!$M$2:$M${LAST_ROW},"WIN")')
        ws2.cell(row=i, column=5, value=f'=COUNTIFS({TD}!${group_col_letter}$2:${group_col_letter}${LAST_ROW},A{i},{TD}!$N$2:$N${LAST_ROW},"YES")')
        ws2.cell(row=i, column=6, value=f'=ROUND(SUMIF({TD}!${group_col_letter}$2:${group_col_letter}${LAST_ROW},A{i},{TD}!$I$2:$I${LAST_ROW}),2)')
        ws2.cell(row=i, column=7, value=f'=ROUND(SUMIF({TD}!${group_col_letter}$2:${group_col_letter}${LAST_ROW},A{i},{TD}!$J$2:$J${LAST_ROW}),2)')
        ws2.cell(row=i, column=8, value=f'=ROUND(SUMIF({TD}!${group_col_letter}$2:${group_col_letter}${LAST_ROW},A{i},{TD}!$K$2:$K${LAST_ROW}),2)')
        ws2.cell(row=i, column=9, value=f'=IFERROR(ROUND(G{i}/ABS(F{i})*100,1),"")')
        ws2.cell(row=i, column=10, value=f'=IF(H{i}>0,"PROFITABLE (net)","LOSING (net)")')
        for c in range(1, 11):
            ws2.cell(row=i, column=c).border = BORDER
    last = 1 + len(group_values)
    ws2.conditional_formatting.add(f"H2:H{last}", CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL))
    ws2.conditional_formatting.add(f"H2:H{last}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL))
    for i, w in enumerate([12, 8, 10, 9, 9, 15, 16, 14, 15, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

build_summary_sheet("Per-Pair Summary", "B", {t["symbol"] for t in usable})
build_summary_sheet("Per-Strategy Summary", "A", {t["strategy"] for t in usable})

# ============================================================
# 3. Forward-observation logs (2026-08-27 infrastructure)
# ============================================================
cost_gate_rows = load_jsonl(os.path.join(DATA_DIR, "cost_gate_decisions.jsonl"))
ws_cg = wb.create_sheet("Cost Gate Log")
style_header(ws_cg, ["Timestamp", "Account", "Strategy", "Pair", "Direction",
                      "Expected Profit (EUR)", "Round-Trip Cost (EUR)", "Ratio", "Decision", "Reason"])
for i, cg in enumerate(cost_gate_rows, 2):
    vals = [cg.get("timestamp", "")[:19], cg.get("account_env"), cg.get("strategy"), cg.get("symbol"),
            cg.get("direction"), cg.get("expected_target_profit_eur"), cg.get("round_trip_cost_eur"),
            cg.get("ratio_actual"), cg.get("decision"), cg.get("reason")]
    for j, v in enumerate(vals, 1):
        c = ws_cg.cell(row=i, column=j, value=v)
        c.border = BORDER
        if j == 9:
            c.fill = RED_FILL if v == "BLOCKED" else GREEN_FILL
for i, w in enumerate([20, 10, 12, 9, 10, 16, 16, 8, 10, 16], 1):
    ws_cg.column_dimensions[get_column_letter(i)].width = w
ws_cg.freeze_panes = "A2"

exposure_rows = load_jsonl(os.path.join(DATA_DIR, "currency_exposure_snapshots.jsonl"))
ws_ex = wb.create_sheet("Exposure Snapshots")
ws_ex["A1"] = f"{len(exposure_rows)} snapshot(s) logged so far (one per run_daily() cycle)"
ws_ex["A1"].font = Font(italic=True, color="666666")
if exposure_rows:
    latest = exposure_rows[-1]
    style_header(ws_ex, ["Currency", "Count Exposure", "Notional (EUR)", "% of Equity"], row=3)
    ranked = sorted(latest.get("notional_exposure_eur", {}).items(), key=lambda kv: -abs(kv[1]))
    for i, (ccy, notional) in enumerate(ranked, 4):
        vals = [ccy, latest.get("count_exposure", {}).get(ccy), round(notional, 2),
                latest.get("pct_of_equity", {}).get(ccy)]
        for j, v in enumerate(vals, 1):
            ws_ex.cell(row=i, column=j, value=v).border = BORDER
    ws_ex.cell(row=2, column=1, value=f"Latest snapshot: {latest.get('timestamp','')[:19]} -- equity {latest.get('equity_eur')} EUR").font = BOLD
for i, w in enumerate([12, 16, 16, 14], 1):
    ws_ex.column_dimensions[get_column_letter(i)].width = w

card_rows = load_jsonl(os.path.join(DATA_DIR, "trade_observation_cards.jsonl"))
entries_by_id = {c["card_id"]: c for c in card_rows if c.get("event") == "entry"}
exits_by_id = {c["card_id"]: c for c in card_rows if c.get("event") == "exit"}
ws_cards = wb.create_sheet("Trade Cards")
style_header(ws_cards, ["Strategy", "Pair", "Direction", "Entry", "ATR", "Current Stop",
                         "Structural Stop", "Hybrid Stop", "Risk (EUR)", "Cost/Edge Ratio",
                         "Status", "Exit Price", "Exit Reason", "Net P&L (EUR)", "R-Multiple",
                         "MAE (EUR)", "MFE (EUR)", "Holding (hrs)"])
for i, (cid, entry) in enumerate(sorted(entries_by_id.items(), key=lambda kv: kv[1]["timestamp"]), 2):
    exit_ = exits_by_id.get(cid)
    vals = [entry.get("strategy"), entry.get("symbol"), entry.get("direction"), entry.get("entry_price"),
            entry.get("atr_at_entry"), entry.get("current_stop"), entry.get("structural_stop"),
            entry.get("hybrid_stop"), entry.get("risk_eur"), entry.get("cost_to_edge_ratio"),
            "closed" if exit_ else "open",
            exit_.get("exit_price") if exit_ else None, exit_.get("exit_reason") if exit_ else None,
            exit_.get("net_pnl_eur") if exit_ else None, exit_.get("r_multiple") if exit_ else None,
            exit_.get("mae_eur") if exit_ else None, exit_.get("mfe_eur") if exit_ else None,
            exit_.get("holding_hours") if exit_ else None]
    for j, v in enumerate(vals, 1):
        ws_cards.cell(row=i, column=j, value=v).border = BORDER
for i, w in enumerate([12, 9, 10, 11, 10, 12, 13, 11, 11, 13, 8, 11, 20, 12, 10, 11, 11, 12], 1):
    ws_cards.column_dimensions[get_column_letter(i)].width = w
ws_cards.freeze_panes = "A2"

# ============================================================
# 4. Verdict Summary (first sheet)
# ============================================================
ws5 = wb.create_sheet("Verdict Summary", 0)
ws5["A1"] = f"ATOS SIM -- Daily Report -- {TODAY}"
ws5["A1"].font = Font(bold=True, size=14)
ws5["A2"] = (f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} PKT -- {len(usable)} trades, live "
             f"commission/prices pulled fresh from Saxo. {len(cost_gate_rows)} cost-gate decisions, "
             f"{len(exposure_rows)} exposure snapshots, {len(entries_by_id)} trade cards logged so far "
             f"in the forward-SIM observation window (started 2026-08-27).")
ws5["A2"].font = Font(italic=True, color="666666")
ws5.merge_cells("A2:H2")
ws5["A2"].alignment = Alignment(wrap_text=True)
labels = [
    ("Total trades analyzed", f'=COUNTA({TD}!B2:B{LAST_ROW})'),
    ("Gross P&L, all trades (EUR)", f'=ROUND(SUM({TD}!I2:I{LAST_ROW}),2)'),
    ("Real commission cost, all trades (EUR)", f'=ROUND(SUM({TD}!J2:J{LAST_ROW}),2)'),
    ("Net P&L, all trades (EUR)", f'=ROUND(SUM({TD}!K2:K{LAST_ROW}),2)'),
    ("Trades that flip win->loss on cost", f'=COUNTIF({TD}!N2:N{LAST_ROW},"YES")'),
    ("Cost-gate decisions logged today", len(cost_gate_rows)),
    ("Cost-gate BLOCKED count", sum(1 for c in cost_gate_rows if c.get("decision") == "BLOCKED")),
]
r0 = 4
for label, val in labels:
    ws5.cell(row=r0, column=1, value=label).font = BOLD
    ws5.cell(row=r0, column=2, value=val)
    r0 += 1
ws5.column_dimensions["A"].width = 44
ws5.column_dimensions["B"].width = 40

out_path = os.path.join(os.path.expanduser("~"), "Downloads", f"sim_daily_report_{TODAY}.xlsx")
wb.save(out_path)
print("Saved:", out_path)
