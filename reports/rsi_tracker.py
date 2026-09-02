"""
reports/rsi_tracker.py  --  RSI strategy tracker, SIM + LIVE
---------------------------------------------------------------
Answers one question: for the core RSI strategy
(forex/strategy_rsi.py, runner key "rsi" -- the one we've run since day 1),
which currency pairs are
*winning* and which are *losing*, across SIM and both real-money accounts
-- with every W/L / WR / PF / net-P&L / avg / best / worst detail.

Naming: "RSI" here always means that one core day-1 strategy. "RSI2" means
the separate SIM-only A/B-test variant "advanced_rsi_master"
(forex/strategy_advanced_rsi_master.py) -- a different strategy, never
confused with the core one.

Scope
  Core strategy only: strategy = "rsi" in pnl_ledger.db. The separate
  variant "advanced_rsi_master" (shown as "RSI2") is NOT core -- pass
  --include-variants to also print it as a side comparison.

  Modules pulled from data/pnl_ledger.db:
    SIM       -> module "forex"
    LIVE-SEK  -> module "forex_live"       (15,000 SEK cap, 17 HIGH_VOLUME pairs)
    LIVE-EUR  -> module "forex_live_eur"   (8,000 EUR cap, 49 CORE pairs)

  "Usable" = a closed trade with a confirmed realized P&L. Administrative
  closes (exit_reason "ledger_dedup*", "reconciled_no_stat*",
  "dedup_stacked_reen*", etc.) have realized_pnl = NULL and are excluded --
  they are bookkeeping, not trades. Same rule as
  reports/module_performance_tracker.py.

  All realized_pnl for forex modules is already net (post-cost) and in the
  ledger base currency, EUR -- see pnl_tracker.log_close's
  gross_pnl_base_override / fx_rate_to_base.

Output
  1. Terminal report (always): headline SIM vs LIVE vs COMBINED, full
     per-pair table, explicit WINNING / LOSING pair lists, open positions.
  2. data/rsi_tracker.xlsx (unless --no-xlsx): one persistent workbook,
     overwritten in place -- Per-Pair (Combined / SIM / LIVE) + hidden
     Trade Detail + Daily/Weekly/Monthly, same live-formula design as the
     other performance trackers (reports/_perf_common.py).

Read-only. Never touches a signal, gate, stop, or order. Run any time.

Usage
    py -3.12 reports/rsi_tracker.py                # terminal + xlsx
    py -3.12 reports/rsi_tracker.py --no-xlsx      # terminal only
    py -3.12 reports/rsi_tracker.py --include-variants
"""
import sys, os, sqlite3, argparse
from collections import defaultdict
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)   # so `import forex.universe` (no torch) works
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH  = os.path.join(DATA_DIR, "pnl_ledger.db")
NOW      = datetime.now().strftime("%Y-%m-%d %H:%M")

CCY = "EUR"
CORE_STRATEGY = "rsi"
VARIANT_STRATEGIES = ("advanced_rsi_master",)

# Display names. "RSI" = the core strategy we've run since day 1
# (forex/strategy_rsi.py). "RSI2" = the separate SIM-only A/B-test variant
# (forex/strategy_advanced_rsi_master.py, key advanced_rsi_master) -- a
# genuinely different strategy, kept clearly distinct from the core one.
STRAT_LABEL = {
    "rsi": "RSI",
    "advanced_rsi_master": "RSI2",
}

# module -> (scope, label). scope is the SIM/LIVE bucket the per-pair
# rollups split on; label is what prints.
MODULES = {
    "forex":          ("SIM",  "SIM"),
    "forex_live":     ("LIVE", "LIVE-SEK"),
    "forex_live_eur": ("LIVE", "LIVE-EUR"),
}

# Forex Grouping tier for each pair -- forex/universe.py has no torch
# dependency (same import pair_group_performance_tracker.py relies on).
# Order = forex_dashboard.py's section order. HIGH_VOLUME (17) is the set
# the LIVE-SEK account trades RSI on; LIVE-EUR trades all 49 CORE
# (HIGH_VOLUME + CORE_STANDARD).
TIER_ORDER = ["High Volume", "Core Standard", "Scandi", "Metals",
              "Exotic Asia", "Exotic Europe", "Exotic High-Yield/Carry",
              "Exotic Latam/Mideast"]
try:
    from forex.universe import get_group as _get_group
    def tier_of(sym):
        try:
            return _get_group(sym)
        except Exception:
            return "?"
except Exception:  # pragma: no cover - universe import should always work
    def tier_of(sym):
        return "?"

# ── ANSI ────────────────────────────────────────────────────────────
if hasattr(sys.stdout, "reconfigure") and sys.stdout.encoding != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
try:
    import ctypes
    k = ctypes.windll.kernel32
    h = k.GetStdHandle(-11)
    m = ctypes.c_ulong()
    k.GetConsoleMode(h, ctypes.byref(m))
    k.SetConsoleMode(h, m.value | 0x4)
except Exception:
    pass

GR, RD, YL, CY, DM, BD, W = (
    "\033[92m", "\033[91m", "\033[93m", "\033[96m", "\033[2m", "\033[1m", "\033[0m",
)


# ── Load ────────────────────────────────────────────────────────────
def load_rows(strategies):
    if not os.path.exists(DB_PATH):
        print(f"ERROR: {DB_PATH} not found.")
        sys.exit(1)
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    ph = ",".join("?" * len(strategies))
    mods = tuple(MODULES)
    mph = ",".join("?" * len(mods))
    rows = c.execute(
        f"SELECT * FROM trades WHERE strategy IN ({ph}) AND module IN ({mph}) ORDER BY id",
        (*strategies, *mods),
    ).fetchall()
    c.close()
    return [dict(r) for r in rows]


def split_rows(rows):
    """(usable_closed, admin_closed, open_) -- usable = closed w/ real P&L."""
    usable, admin, open_ = [], [], []
    for r in rows:
        if r["status"] == "open":
            open_.append(r)
        elif r["status"] == "closed" and r["realized_pnl"] is not None:
            usable.append(r)
        else:
            admin.append(r)
    return usable, admin, open_


# ── Stats ───────────────────────────────────────────────────────────
def agg(trades):
    n = len(trades)
    wins = [t for t in trades if t["realized_pnl"] > 0]
    losses = [t for t in trades if t["realized_pnl"] < 0]
    scratch = n - len(wins) - len(losses)
    gp = sum(t["realized_pnl"] for t in wins)
    gl = sum(t["realized_pnl"] for t in losses)  # <= 0
    net = sum(t["realized_pnl"] for t in trades)
    pf = (gp / abs(gl)) if gl < 0 else (float("inf") if gp > 0 else None)
    pnls = [t["realized_pnl"] for t in trades]
    return {
        "n": n, "wins": len(wins), "losses": len(losses), "scratch": scratch,
        "wr": (len(wins) / n * 100) if n else 0.0,
        "gross_profit": gp, "gross_loss": gl, "net": net,
        "pf": pf,
        "avg": (net / n) if n else 0.0,
        "best": max(pnls) if pnls else 0.0,
        "worst": min(pnls) if pnls else 0.0,
        "last_close": max((t["timestamp_close"] or "" for t in trades), default="")[:10],
    }


def pf_str(pf):
    if pf is None:
        return "--"
    if pf == float("inf"):
        return "inf"
    return f"{pf:.2f}"


def fmt_money(v):
    return f"{v:+,.2f}"


# ── Terminal report ─────────────────────────────────────────────────
def print_report(usable, admin, open_, include_variants, all_rows):
    core = [t for t in usable if t["strategy"] == CORE_STRATEGY]
    by_scope = defaultdict(list)
    for t in core:
        by_scope[MODULES[t["module"]][0]].append(t)

    line = "=" * 92
    print(f"\n{BD}{CY}{line}{W}")
    print(f"{BD}{CY}  RSI STRATEGY TRACKER  --  SIM + LIVE   ({CCY}, net of costs){W}")
    print(f"{BD}{CY}  generated {NOW} PKT   ·   core strategy: forex/strategy_rsi.py  (\"rsi\"){W}")
    print(f"{BD}{CY}{line}{W}")
    print(f"  {DM}Modules: SIM=forex · LIVE-SEK=forex_live · LIVE-EUR=forex_live_eur{W}")
    print(f"  {DM}Usable = closed trade with a confirmed realized P&L. "
          f"{len(admin)} admin/dedup close(s) excluded.{W}")

    # ── headline ────────────────────────────────────────────────────
    print(f"\n{BD}  HEADLINE — core RSI{W}")
    hdr = (f"  {'scope':<10} {'trades':>6} {'W':>4} {'L':>4} {'WR%':>7} "
           f"{'net P&L':>13} {'PF':>6} {'avg':>10} {'best':>11} {'worst':>11}")
    print(f"{DM}{hdr}{W}")
    print(f"  {DM}{'-'*88}{W}")

    def prow(label, s):
        col = GR if s["net"] >= 0 else RD
        print(f"  {BD}{label:<10}{W} {s['n']:>6} {s['wins']:>4} {s['losses']:>4} "
              f"{s['wr']:>6.1f}% {col}{fmt_money(s['net']):>13}{W} {pf_str(s['pf']):>6} "
              f"{fmt_money(s['avg']):>10} {GR}{fmt_money(s['best']):>11}{W} {RD}{fmt_money(s['worst']):>11}{W}")

    prow("SIM", agg(by_scope["SIM"]))
    prow("LIVE", agg(by_scope["LIVE"]))
    print(f"  {DM}{'-'*88}{W}")
    prow("COMBINED", agg(core))

    op = defaultdict(int)
    for r in open_:
        if r["strategy"] == CORE_STRATEGY:
            op[MODULES[r["module"]][1]] += 1
    if op:
        print(f"  {DM}open RSI positions: " + " · ".join(f"{k} {v}" for k, v in sorted(op.items())) + f"{W}")

    # ── tier scorecard ─────────────────────────────────────────────
    # Forex Grouping tier -- the LIVE-SEK account trades RSI ONLY on the 17
    # High Volume pairs; LIVE-EUR trades all 49 CORE (High Volume + Core
    # Standard). So "is RSI better on Core Standard than on High Volume?"
    # is directly a "which live account should keep running RSI?" question.
    by_tier = defaultdict(list)
    for t in core:
        by_tier[tier_of(t["symbol"])].append(t)
    print(f"\n{BD}  BY TIER — core RSI (Forex Grouping){W}")
    th = (f"  {'tier':<24} {'trades':>6} {'W':>3} {'L':>3} {'WR%':>7} "
          f"{'net P&L':>13} {'PF':>7} {'avg/trade':>11}")
    print(f"{DM}{th}{W}")
    print(f"  {DM}{'-'*80}{W}")
    tier_keys = [k for k in TIER_ORDER if k in by_tier] + \
                sorted(k for k in by_tier if k not in TIER_ORDER)
    for tk in tier_keys:
        s = agg(by_tier[tk])
        col = GR if s["net"] >= 0 else RD
        note = ""
        if tk == "High Volume":
            note = f"  {DM}<- LIVE-SEK trades RSI only here{W}"
        elif tk == "Core Standard":
            note = f"  {DM}<- LIVE-EUR also trades these{W}"
        print(f"  {BD}{tk:<24}{W} {s['n']:>6} {s['wins']:>3} {s['losses']:>3} "
              f"{s['wr']:>6.1f}% {col}{fmt_money(s['net']):>13}{W} {pf_str(s['pf']):>7} "
              f"{fmt_money(s['avg']):>11}{note}")

    # ── per-pair (combined) ─────────────────────────────────────────
    by_pair = defaultdict(list)
    for t in core:
        by_pair[t["symbol"]].append(t)
    pair_stats = {sym: agg(ts) for sym, ts in by_pair.items()}
    ordered = sorted(pair_stats, key=lambda s: pair_stats[s]["net"], reverse=True)

    print(f"\n{BD}  PER-PAIR — core RSI, SIM+LIVE combined  ({len(ordered)} pairs traded){W}")
    ph = (f"  {'pair':<9} {'src':<10} {'trades':>6} {'W':>3} {'L':>3} {'WR%':>7} "
          f"{'net P&L':>13} {'PF':>6} {'avg':>10} {'best':>10} {'worst':>10} {'last':>11}")
    print(f"{DM}{ph}{W}")
    print(f"  {DM}{'-'*108}{W}")
    for sym in ordered:
        s = pair_stats[sym]
        srcs = sorted({MODULES[t["module"]][1] for t in by_pair[sym]})
        src = "+".join(x.replace("LIVE-", "L-").replace("SIM", "SIM") for x in srcs)
        col = GR if s["net"] >= 0 else RD
        print(f"  {BD}{sym:<9}{W} {DM}{src:<10}{W} {s['n']:>6} {s['wins']:>3} {s['losses']:>3} "
              f"{s['wr']:>6.1f}% {col}{fmt_money(s['net']):>13}{W} {pf_str(s['pf']):>6} "
              f"{fmt_money(s['avg']):>10} {fmt_money(s['best']):>10} {fmt_money(s['worst']):>10} "
              f"{DM}{s['last_close']:>11}{W}")

    winners = [s for s in ordered if pair_stats[s]["net"] > 0]
    losers = [s for s in ordered if pair_stats[s]["net"] < 0]
    flat = [s for s in ordered if pair_stats[s]["net"] == 0]

    print(f"\n{BD}{GR}  WINNING PAIRS — core RSI is net positive here  ({len(winners)}){W}")
    if winners:
        tot = sum(pair_stats[s]["net"] for s in winners)
        for s in winners:
            st = pair_stats[s]
            print(f"    {GR}+{W} {BD}{s:<9}{W} {fmt_money(st['net']):>12} {CCY}   "
                  f"{st['wins']}W/{st['losses']}L  WR {st['wr']:.0f}%  PF {pf_str(st['pf'])}")
        print(f"    {DM}{'':<11}{'-'*24}{W}")
        print(f"    {GR}={W} {BD}{'subtotal':<9}{W} {GR}{fmt_money(tot):>12} {CCY}{W}")
    else:
        print(f"    {DM}(none){W}")

    print(f"\n{BD}{RD}  LOSING PAIRS — core RSI is net negative here  ({len(losers)}){W}")
    if losers:
        tot = sum(pair_stats[s]["net"] for s in losers)
        for s in losers:
            st = pair_stats[s]
            print(f"    {RD}-{W} {BD}{s:<9}{W} {fmt_money(st['net']):>12} {CCY}   "
                  f"{st['wins']}W/{st['losses']}L  WR {st['wr']:.0f}%  PF {pf_str(st['pf'])}")
        print(f"    {DM}{'':<11}{'-'*24}{W}")
        print(f"    {RD}={W} {BD}{'subtotal':<9}{W} {RD}{fmt_money(tot):>12} {CCY}{W}")
    if flat:
        print(f"  {DM}flat / scratch: {', '.join(flat)}{W}")

    # ── open positions detail ──────────────────────────────────────
    core_open = [r for r in open_ if r["strategy"] == CORE_STRATEGY]
    if core_open:
        print(f"\n{BD}  OPEN RSI POSITIONS  ({len(core_open)}){W}")
        print(f"  {DM}{'pair':<9} {'src':<9} {'dir':<5} {'units':>12} {'entry':>14} {'opened':>17}{W}")
        for r in sorted(core_open, key=lambda x: (MODULES[x['module']][1], x['symbol'])):
            print(f"  {r['symbol']:<9} {DM}{MODULES[r['module']][1]:<9}{W} {r['direction']:<5} "
                  f"{r['quantity']:>12,.0f} {r['entry_price']:>14,.5f} "
                  f"{DM}{(r['timestamp_open'] or '')[:16]:>17}{W}")

    # ── optional variant comparison ────────────────────────────────
    if include_variants:
        var = [t for t in usable if t["strategy"] in VARIANT_STRATEGIES]
        print(f"\n{BD}{YL}  SIDE COMPARISON — RSI2 = advanced_rsi_master, SIM-only A/B variant "
              f"(a SEPARATE strategy, NOT the core RSI){W}")
        if var:
            vb = defaultdict(list)
            for t in var:
                vb[t["strategy"]].append(t)
            for strat, ts in vb.items():
                s = agg(ts)
                col = GR if s["net"] >= 0 else RD
                print(f"    {STRAT_LABEL.get(strat, strat):<8} {DM}({strat}){W}  {s['n']:>3} trades  "
                      f"{s['wins']}W/{s['losses']}L  WR {s['wr']:.0f}%  PF {pf_str(s['pf'])}  "
                      f"net {col}{fmt_money(s['net'])} {CCY}{W}")
        else:
            print(f"    {DM}(no usable closed trades){W}")

    print(f"\n{BD}{CY}{line}{W}")
    print(f"  {DM}small-sample caveat: core RSI has {len(core)} usable closed trades total "
          f"-- per-pair reads are directional, not yet statistically settled.{W}")
    print(f"{BD}{CY}{line}{W}\n")

    return pair_stats, ordered, by_pair


# ── Excel ───────────────────────────────────────────────────────────
def build_xlsx(usable, all_rows):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule
    except ModuleNotFoundError:
        print(f"{YL}[rsi_tracker] openpyxl not available on this interpreter -- "
              f"skipping data/rsi_tracker.xlsx (run with py -3.12). Terminal report above is complete.{W}")
        return
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from _perf_common import (parse_close_date, day_key, week_key, month_key,
                              sorted_distinct_period_keys)

    def metric_formulas(ws, row_i, crit, last_row, start_col=2):
        """Trades / Wins / WR% / Gross / Commission / Net / Profit Factor as
        7 live formulas. `crit` is one or more "range,criteria" pairs (a
        COUNTIFS/SUMIFS-style fragment) -- unlike _perf_common's helper this
        always uses the *IFS forms, so a compound crit (e.g. pair AND
        source) is valid."""
        from openpyxl.utils import get_column_letter as _L
        c = start_col
        trades, wins, wr, gross, comm, net, pf = (_L(c + i) for i in range(7))
        g = f"{TD}!$G$2:$G${last_row}"
        h = f"{TD}!$H$2:$H${last_row}"
        i_ = f"{TD}!$I$2:$I${last_row}"
        j = f"{TD}!$J$2:$J${last_row}"
        ws.cell(row=row_i, column=c,     value=f'=COUNTIFS({crit})')
        ws.cell(row=row_i, column=c + 1, value=f'=COUNTIFS({crit},{j},"WIN")')
        ws.cell(row=row_i, column=c + 2, value=f'=IFERROR(ROUND({wins}{row_i}/{trades}{row_i}*100,1),"")')
        ws.cell(row=row_i, column=c + 3, value=f'=ROUND(SUMIFS({g},{crit}),2)')
        ws.cell(row=row_i, column=c + 4, value=f'=ROUND(SUMIFS({h},{crit}),2)')
        ws.cell(row=row_i, column=c + 5, value=f'=ROUND(SUMIFS({i_},{crit}),2)')
        win_sum = f'SUMIFS({i_},{crit},{i_},">0")'
        loss_sum = f'SUMIFS({i_},{crit},{i_},"<0")'
        ws.cell(row=row_i, column=c + 6,
                value=f'=IFERROR(ROUND({win_sum}/ABS({loss_sum}),2),'
                      f'IF({trades}{row_i}=0,"",">0 (no losers)"))')

    HEADER_FILL = PatternFill("solid", fgColor="1F2937")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    RED_FILL = PatternFill("solid", fgColor="FCE4E4")
    GREEN_FILL = PatternFill("solid", fgColor="E4F7E4")
    BOLD = Font(bold=True)
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def style_header(ws, headers, row=1):
        for i, h in enumerate(headers, 1):
            cc = ws.cell(row=row, column=i, value=h)
            cc.fill, cc.font = HEADER_FILL, HEADER_FONT
            cc.alignment = Alignment(wrap_text=True, vertical="center")
            cc.border = BORDER

    core = [t for t in usable if t["strategy"] == CORE_STRATEGY]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trade Detail"
    # Canonical _perf_common layout: summary-sheet formulas hardcode
    # G/H/I = Gross/Commission/Net, J = Net Result, K/L/M = Close/Week/Month.
    # Column C is the one free "grouping" slot -- used here for Source
    # (SIM / LIVE-SEK / LIVE-EUR); the LIVE per-pair sheet wildcard-matches
    # "LIVE*" on it.
    # Column N (past the canonical A..M block) = Forex Grouping tier -- the
    # Per-Tier sheet keys on it. G/H/I/J/K/L/M keep their canonical meaning.
    style_header(ws, ["Strategy", "Symbol", "Source", "Direction", "Units", "Status",
                      f"Gross P&L ({CCY})", f"Commission ({CCY})", f"Net P&L ({CCY})", "Net Result",
                      "Close Date", "Week", "Month", "Tier"])
    r = 2
    for t in sorted(core, key=lambda x: x["timestamp_close"] or ""):
        net = round(t["realized_pnl"], 2)
        comm = round(t["commission"] or 0.0, 2)
        res = "WIN" if net > 0 else "LOSS" if net < 0 else "SCRATCH"
        d = parse_close_date(t["timestamp_close"])
        _scope, srclabel = MODULES[t["module"]]
        vals = [STRAT_LABEL.get(t["strategy"], t["strategy"]), t["symbol"], srclabel, t["direction"],
                t["quantity"], "closed", round(net + comm, 2), comm, net, res,
                day_key(d) if d else "", week_key(d) if d else "", month_key(d) if d else "",
                tier_of(t["symbol"])]
        for i, v in enumerate(vals, 1):
            cc = ws.cell(row=r, column=i, value=v)
            cc.border = BORDER
            if i == 10:
                cc.fill = GREEN_FILL if res == "WIN" else RED_FILL if res == "LOSS" else None
        r += 1
    LAST = r - 1
    for i, wdt in enumerate([18, 10, 10, 9, 12, 8, 14, 13, 12, 10, 12, 10, 9, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A2"
    ws.sheet_state = "hidden"
    TD = "'Trade Detail'"

    HEADERS = ["Key", "Trades", "Wins (net)", "WR %", f"Gross P&L ({CCY})",
               f"Commission ({CCY})", f"Net P&L ({CCY})", "Profit Factor"]
    WIDTHS = [22, 9, 11, 8, 16, 16, 15, 13]

    usable_recs = [{"status": "closed", "timestamp_close": t["timestamp_close"]} for t in core]

    def key_sheet(title, td_col, keys, index=None, extra_crit=None):
        w2 = wb.create_sheet(title, index) if index is not None else wb.create_sheet(title)
        w2["A1"] = f"RSI Tracker -- {title} -- updated {NOW} PKT ({CCY}, net)"
        w2["A1"].font = Font(bold=True, size=13)
        w2.merge_cells("A1:H1")
        style_header(w2, HEADERS, row=3)
        rr = 4
        for key in keys:
            w2.cell(row=rr, column=1, value=key).font = BOLD
            crit = f'{TD}!${td_col}$2:${td_col}${LAST},"{key}"'
            if extra_crit:
                crit = crit + "," + extra_crit
            metric_formulas(w2, rr, crit, LAST, start_col=2)
            for cc in range(1, 9):
                w2.cell(row=rr, column=cc).border = BORDER
            rr += 1
        last = rr - 1
        if last >= 4:
            w2.conditional_formatting.add(f"G4:G{last}", CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL))
            w2.conditional_formatting.add(f"G4:G{last}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL))
        else:
            w2.cell(row=4, column=1, value="No usable closed trades yet.").font = Font(italic=True, color="999999")
        for i, wdt in enumerate(WIDTHS, 1):
            w2.column_dimensions[get_column_letter(i)].width = wdt
        w2.freeze_panes = "A4"
        return w2

    pairs = sorted({t["symbol"] for t in core})
    key_sheet("Per-Pair (Combined)", "B", pairs, index=0)
    key_sheet("Per-Pair SIM", "B", sorted({t["symbol"] for t in core if MODULES[t["module"]][0] == "SIM"}),
              extra_crit=f'{TD}!$C$2:$C${LAST},"SIM"')
    key_sheet("Per-Pair LIVE", "B", sorted({t["symbol"] for t in core if MODULES[t["module"]][0] == "LIVE"}),
              extra_crit=f'{TD}!$C$2:$C${LAST},"LIVE*"')
    key_sheet("Per-Source", "C", [m[1] for m in MODULES.values()])
    tiers_present = [k for k in TIER_ORDER if any(tier_of(t["symbol"]) == k for t in core)]
    key_sheet("Per-Tier", "N", tiers_present, index=1)
    key_sheet("Daily", "K", sorted_distinct_period_keys(usable_recs, day_key))
    key_sheet("Weekly", "L", sorted_distinct_period_keys(usable_recs, week_key))
    key_sheet("Monthly", "M", sorted_distinct_period_keys(usable_recs, month_key))

    out = os.path.join(DATA_DIR, "rsi_tracker.xlsx")
    wb.save(out)
    print(f"  {GR}saved{W} {out}  ({LAST - 1} usable core-RSI trades)\n")


def main():
    ap = argparse.ArgumentParser(description="RSI strategy tracker -- SIM + LIVE")
    ap.add_argument("--no-xlsx", action="store_true", help="terminal report only")
    ap.add_argument("--include-variants", action="store_true",
                    help="also show SIM-only advanced_rsi_master as a side comparison")
    args = ap.parse_args()

    strategies = [CORE_STRATEGY] + (list(VARIANT_STRATEGIES) if args.include_variants else [])
    all_rows = load_rows(strategies)
    usable, admin, open_ = split_rows(all_rows)
    print_report(usable, admin, open_, args.include_variants, all_rows)
    if not args.no_xlsx:
        build_xlsx(usable, all_rows)


if __name__ == "__main__":
    main()
