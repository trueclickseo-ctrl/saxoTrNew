"""
Regression tests -- 2026-08-28 daily PF/PNL/WR performance trackers for
ETF, Stocks, Futures (+ forex's Daily/Weekly/Monthly sheets).

Explicit user request: "also make advance Excel tracker for ETF,
Stocks, Futures -- keep track daily, weekly, monthly like we have
advance Excel tracker for Forex."

Unlike forex's tracker (reports/pair_group_performance_tracker.py,
which needs live EUR re-pricing across many quote currencies + a Forex
Grouping tier breakdown), the other 3 modules are single-currency and
pnl_ledger.db's trades table already stores each closed trade's real
dealt realized_pnl (net)/commission -- so reports/module_performance_
tracker.py reads directly from the DB, no live re-pricing, no torch
dependency, single self-contained script.

NOTE: run this file with `py -3.12`, not the normal project `python` --
it imports openpyxl at module level, same as every report script here
(the normal project Python has forex.runner/torch but not openpyxl on
this machine; py -3.12 has openpyxl but not torch -- see
reports/daily_sim_report.py's docstring for the full rationale):

    py -3.12 test_2026_08_28_module_performance_trackers.py
"""

import os
import subprocess
import sys

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

GREEN, RED, YELLOW, CYAN, RESET, BOLD = (
    "\033[92m", "\033[91m", "\033[93m", "\033[96m", "\033[0m", "\033[1m"
)
_results = []


def _run(name, fn):
    try:
        result = fn()
        if result is None:
            result = True
        _results.append((name, bool(result), None))
    except Exception as e:
        _results.append((name, False, f"{type(e).__name__}: {e}"))


def section(title):
    print(f"\n{BOLD}{CYAN}{'-'*70}{RESET}")
    print(f"{BOLD}{CYAN}  {title}{RESET}")
    print(f"{BOLD}{CYAN}{'-'*70}{RESET}")


def _py312():
    # Same interpreter every .bat/report script uses for the openpyxl phase.
    return "py"


def _run_module_tracker(module, display, ccy):
    return subprocess.run(
        [_py312(), "-3.12", os.path.join(BASE_DIR, "reports", "module_performance_tracker.py"), module, display, ccy],
        cwd=BASE_DIR, capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=60,
    )


# ═══════════════════════════════════════════════════════════════════════
section("1. reports/module_performance_tracker.py -- futures/etf/stock")
# ═══════════════════════════════════════════════════════════════════════

MODULES = [("futures", "Futures", "USD"), ("etf", "ETF", "USD"), ("stock", "Stocks", "SEK")]


def test_all_three_module_trackers_run_cleanly():
    for module, display, ccy in MODULES:
        proc = _run_module_tracker(module, display, ccy)
        assert proc.returncode == 0, f"{module} tracker failed: {proc.stderr}"
_run("module_performance_tracker.py runs cleanly for futures/etf/stock",
     test_all_three_module_trackers_run_cleanly)


def test_all_three_workbooks_exist_with_expected_sheets():
    expected_sheets = {"Per-Strategy Performance", "Per-Symbol Performance", "Trade Detail",
                        "Daily Performance", "Weekly Performance", "Monthly Performance"}
    for module, _, _ in MODULES:
        path = os.path.join(BASE_DIR, "data", f"{module}_performance_tracker.xlsx")
        assert os.path.exists(path), f"expected {path} to exist"
        wb = openpyxl.load_workbook(path)
        assert set(wb.sheetnames) == expected_sheets, (
            f"{module}: expected exactly {expected_sheets}, got {set(wb.sheetnames)}"
        )
_run("Each module's workbook exists with exactly the 6 expected sheets",
     test_all_three_workbooks_exist_with_expected_sheets)


def test_trade_detail_is_hidden_and_summary_sheets_use_live_formulas():
    for module, _, _ in MODULES:
        path = os.path.join(BASE_DIR, "data", f"{module}_performance_tracker.xlsx")
        wb = openpyxl.load_workbook(path)
        assert wb["Trade Detail"].sheet_state == "hidden", f"{module}: Trade Detail should be hidden"
        ws = wb["Per-Strategy Performance"]
        # Row 4 col B should hold a live formula (starts with '='), not a
        # pre-computed static number -- this is what makes the sheet
        # self-update if Trade Detail is ever regenerated independently.
        cell = ws.cell(row=4, column=2).value
        if cell is not None:
            assert isinstance(cell, str) and cell.startswith("="), (
                f"{module}: expected a live formula in Per-Strategy Performance!B4, got {cell!r}"
            )
_run("Trade Detail sheet is hidden; summary sheets use live COUNTIF/SUMIF formulas, not static values",
     test_trade_detail_is_hidden_and_summary_sheets_use_live_formulas)


def test_futures_per_symbol_shows_real_markets_not_missing():
    # Cross-check against pnl_tracker.get_strategy_summary(), which the
    # dashboard's own Markets-column fix (same day) already validated
    # returns real tickers for futures. pnl_tracker's DISTINCT-symbol
    # query doesn't filter out closed-but-null-realized_pnl rows (id 60,
    # CL -- a documented, deliberately-unresolved historical record, see
    # reports/module_performance_tracker.py's `unresolved` handling), so
    # this tracker's Per-Symbol sheet must still show a row for it
    # (explicitly marked "P&L UNKNOWN"), not silently omit it -- the two
    # views' symbol SETS should match even though CL's row has no
    # computed WR/PF.
    import pnl_tracker
    wb = openpyxl.load_workbook(os.path.join(BASE_DIR, "data", "futures_performance_tracker.xlsx"))
    ws = wb["Per-Symbol Performance"]
    rows_by_symbol = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0]:
            rows_by_symbol[row[0]] = row
    strat_rows = pnl_tracker.get_strategy_summary("futures")
    db_symbols = set()
    for r in strat_rows:
        db_symbols.update(r.get("symbols", []))
    assert set(rows_by_symbol) == db_symbols, (
        f"Per-Symbol Performance sheet's symbols {set(rows_by_symbol)} should exactly match "
        f"pnl_tracker's real traded symbols {db_symbols}"
    )
    assert "CL" in rows_by_symbol, "expected a CL row (closed trade, unknown P&L) to still be shown"
    assert "UNKNOWN" in str(rows_by_symbol["CL"][1]), (
        f"expected CL's row to be explicitly marked P&L UNKNOWN, got {rows_by_symbol['CL']}"
    )
_run("futures_performance_tracker.xlsx's Per-Symbol sheet lists every real traded market, incl. the 1 unresolved (CL) as explicitly P&L UNKNOWN",
     test_futures_per_symbol_shows_real_markets_not_missing)


# ═══════════════════════════════════════════════════════════════════════
section("2. forex's tracker -- new Daily/Weekly/Monthly sheets")
# ═══════════════════════════════════════════════════════════════════════

def test_forex_tracker_has_time_bucketed_sheets():
    path = os.path.join(BASE_DIR, "data", "forex_performance_tracker.xlsx")
    assert os.path.exists(path), "expected data/forex_performance_tracker.xlsx to exist (run its .bat first)"
    wb = openpyxl.load_workbook(path)
    for sheet in ("Daily Performance", "Weekly Performance", "Monthly Performance"):
        assert sheet in wb.sheetnames, f"expected a {sheet!r} sheet in forex's tracker"
_run("forex_performance_tracker.xlsx has Daily/Weekly/Monthly Performance sheets",
     test_forex_tracker_has_time_bucketed_sheets)


def test_forex_daily_sheet_periods_are_real_iso_dates():
    import datetime
    path = os.path.join(BASE_DIR, "data", "forex_performance_tracker.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["Daily Performance"]
    periods = [row[0] for row in ws.iter_rows(min_row=4, max_col=1, values_only=True) if row[0]]
    assert periods, "expected at least one daily period row (forex has closed trades)"
    for p in periods:
        datetime.date.fromisoformat(p)  # raises if not a real ISO date
    assert periods == sorted(periods), "expected daily periods sorted oldest-first"
_run("forex's Daily Performance sheet lists real, chronologically-sorted ISO dates",
     test_forex_daily_sheet_periods_are_real_iso_dates)


print(f"\n{BOLD}{'='*70}{RESET}")
passed = sum(1 for _, ok, _ in _results)
failed = [(n, e) for n, ok, e in _results if not ok]
for name, ok, err in _results:
    icon = f"{GREEN}PASS{RESET}" if ok else f"{RED}FAIL{RESET}"
    print(f"  [{icon}] {name}")
    if err:
        print(f"         {YELLOW}{err}{RESET}")
print(f"{BOLD}{'='*70}{RESET}")
if failed:
    print(f"{RED}{BOLD}  {len(failed)} / {len(_results)} TESTS FAILED{RESET}")
    sys.exit(1)
else:
    print(f"{GREEN}{BOLD}  ALL {len(_results)} TESTS PASSED{RESET}")
    sys.exit(0)
